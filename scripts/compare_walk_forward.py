#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Walk-forward 实验对比脚本

功能：
- 无参时自动扫描 data/walk_forward/raw/ 与 data/walk_forward/batches/*/raw/ 两类来源
- 按 wf_run_id 分组，跨 split 聚合各项指标
- 生成对比表格（行=实验，列=聚合指标+训练参数）
- 输出到 Excel 文件

使用示例：
    python scripts/compare_walk_forward.py
    python scripts/compare_walk_forward.py --data-root ./data
    python scripts/compare_walk_forward.py --raw-dir ./data/walk_forward/raw --output ./data/walk_forward/wf_comparison.csv
"""

import argparse
from bisect import bisect_right
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Optional

project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

import pandas as pd
import numpy as np
from loguru import logger
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.hyperlink import Hyperlink
from dateutil.relativedelta import relativedelta

from src.lazybull.data import DataLoader, Storage
from src.lazybull.common.config import get_data_root
from src.lazybull.common.logger import setup_logger


SUMMARY_CSV_DTYPE = {
    "wf_run_id": str,
    "batch_run_id": str,
    "batch_period_label": str,
    "split_count": str,
    "final_date": str,
    "wf_start_date": str,
    "wf_end_date": str,
}


_TRADE_CAL_CACHE: dict[str, Optional[pd.DataFrame]] = {}
_TRADE_CAL_OPEN_DATES_CACHE: dict[int, tuple[list[str], dict[str, int]]] = {}
_BT_REBALANCE_FREQ_MAX = 60


# ---------------------------------------------------------------------------
# 输出列名：英文内部键 → 中文列名（用于最终 CSV 输出）
# ---------------------------------------------------------------------------
COL_NAMES = {
    # 标识
    "wf_run_id": "运行ID",
    "KEY_说明": "重点说明",
    "KEY_Top20_list": "重点Top20最新名单",
    "KEY_Top30_list": "重点Top30最新名单",
    "key_top20_hit_rate_mean": "重点Top20命中率均值",
    "key_top20_avg_return_median_mean": "重点Top20收益中位数均值",
    "key_top30_hit_rate_mean": "重点Top30命中率均值",
    "key_top30_avg_return_median_mean": "重点Top30收益中位数均值",
    "batch_run_id": "批次ID",
    "batch_period_label": "批次时间段",
    "split_count": "切分数量",
    "final_date": "最终日期",
    # OOS 性能
    "n_splits": "切分数",
    "model_version_range": "模型版本范围",
    "oos_rankic_ir_mean": "OOS_RankIC_IR均值",
    "oos_rankic_ir_std": "OOS_RankIC_IR标准差",
    "oos_cross_split_ir": "跨切分IR",
    "daily_rankic_mean": "RankIC均值",
    "icir": "ICIR",
    "selection_monotonicity": "分层单调性(近似)",
    "oos_rankic_ir_trend": "RankIC_IR趋势(近-早)",
    "oos_top30_median_mean": "Top30中位收益均值",
    "oos_top30_win_rate": "Top30胜率",
    "oos_top30_worst_median": "Top30最差中位收益",
    "oos_top30_skew_score_mean": "Top30偏斜度均值",
    "oos_top30_lift_mean": "Top30超额均值",
    "oos_top100_median_mean": "Top100中位收益均值",
    "oos_top100_win_rate": "Top100胜率",
    "oos_top300_median_mean": "Top300中位收益均值",
    "oos_top300_win_rate": "Top300胜率",
    # OOS 回测
    "chain_cagr": "全周期CAGR",
    "chain_total_return": "全周期总收益",
    "chain_max_drawdown": "全周期链式最大回撤",
    "chain_sharpe": "全周期链式夏普",
    "chain_trading_days": "全周期链式交易日数",
    "bt_annual_return_mean": "回测年化收益均值",
    "bt_sharpe_mean": "回测夏普均值",
    "bt_max_drawdown_worst": "回测最大回撤(最差)",
    "bt_calmar_mean": "回测Calmar均值",
    "bt_win_rate": "回测胜率",
    "bt_total_return_mean": "回测总收益均值",
    "bt_volatility_mean": "回测波动率均值",
    "bt_signal_confidence_block_rate_mean": "门控持币率均值",
    "bt_signal_confidence_avg_exposure_mean": "门控平均仓位",
    "bt_signal_confidence_avg_score_mean": "门控平均置信度",
    "bt_rebalance_freq": "回测调仓频率",
    "bt_initial_capital": "回测初始资金",
    "bt_sell_timing": "回测卖出时机",
    "bt_exclude_st": "回测排除ST",
    "bt_min_list_days": "回测最少上市天数",
    "bt_max_weight_per_stock": "回测单股最大权重",
    "bt_max_per_industry": "回测单行业最大持仓数",
    "bt_stop_loss_enabled": "回测止损",
    "bt_stop_loss_drawdown_pct": "回测回撤止损%",
    "bt_stop_loss_consecutive_limit_down": "回测连续跌停止损",
    "bt_equity_curve_enabled": "回测ECT",
    "bt_equity_curve_drawdown_thresholds": "回测ECT回撤阈值",
    "bt_equity_curve_exposure_levels": "回测ECT仓位系数",
    "bt_equity_curve_ma_short": "回测ECT短均线",
    "bt_equity_curve_ma_long": "回测ECT长均线",
    "bt_equity_curve_recovery_mode": "回测ECT恢复模式",
    "bt_equity_curve_recovery_step": "回测ECT恢复步长",
    "bt_equity_curve_recovery_delay_periods": "回测ECT恢复等待",
    # 训练质量
    "val_rankic_ir_mean": "验证集RankIC_IR均值",
    "train_val_ir_gap": "验证_OOS_IR差距",
    "best_iter_mean": "最佳迭代均值",
    "best_iter_min": "最佳迭代最小值",
    "best_iter_max": "最佳迭代最大值",
    "best_iter_std": "最佳迭代标准差",
    # 跨时间段稳定性
    "period_count": "时间段数",
    "period_labels": "时间段列表",
    "run_id_list": "运行ID列表",
    "score_mean": "综合得分均值",
    "score_std": "综合得分标准差",
    "score_min": "综合得分最差",
    "score_max": "综合得分最佳",
    "chain_cagr_mean": "跨时间段CAGR均值",
    "chain_cagr_std": "跨时间段CAGR标准差",
    "chain_cagr_min": "跨时间段CAGR最差",
    "chain_max_drawdown_mean": "跨时间段回撤均值",
    "chain_max_drawdown_worst": "跨时间段回撤最差",
    "oos_cross_split_ir_mean": "跨时间段跨切分IR均值",
    "oos_cross_split_ir_std": "跨时间段跨切分IR标准差",
    "bt_win_rate_mean": "跨时间段回测胜率均值",
    "bt_win_rate_min": "跨时间段回测胜率最差",
    "chain_sharpe_mean": "跨时间段夏普均值",
    "chain_sharpe_std": "跨时间段夏普标准差",
    "stability_score": "时间段稳定性分",
    "seed_alpha_median": "模型Alpha分中位数",
    # 训练参数
    "split_count": "切分数量",
    "final_date": "最终日期",
    "wf_start_date": "WF起始日期",
    "wf_end_date": "WF结束日期",
    "step": "滚动频率",
    "train_window_years": "训练窗口年数",
    "test_window_months": "测试窗口月数",
    "val_ratio": "验证集比例",
    "label_column": "标签列",
    "task": "任务类型",
    "label_transform": "标签变换",
    "n_estimators": "树数量",
    "max_depth": "最大深度",
    "num_leaves": "LGB叶子数",
    "learning_rate": "学习率",
    "subsample": "样本采样比",
    "colsample_bytree": "特征采样比",
    "min_child_weight": "最小叶节点权重",
    "gamma": "gamma",
    "reg_alpha": "L1正则",
    "reg_lambda": "L2正则",
    "early_stopping_rounds": "早停轮数",
    "early_stopping_metric": "早停指标",
    "adaptive_best_iter_retrain": "自适应候选重训",
    "adaptive_low_iter_max_retries": "自适应低迭代重试上限",
    "ensemble_seeds": "多种子bagging种子",
    "ensemble_seed_keep_top_ratio": "多种子保留比例",
    "ensemble_seed_keep_min_models": "多种子最少保留模型数",
    "rank_weight_enabled": "rank权重启用",
    "rank_weight_topk": "rank权重TopK",
    "rank_weight": "rank权重值",
    "rank_weight_topk_weight_mode": "rank权重TopK模式",
    "time_decay_half_life": "时间衰减半衰期",
    "freshness_strategy": "freshness策略",
    "event_freshness_half_life_days": "事件freshness半衰期天数",
    "objective": "目标函数",
    "algorithm": "算法",
    "enable_fundamental": "基本面因子",
    "enable_alt": "另类因子",
    "enable_margin": "融资融券因子",
    "enable_cyq": "筹码胜率因子",
    "enable_fund": "基金持仓因子",
    "enable_express": "业绩快报因子",
    "feature_stability_filter": "特征稳定性筛选",
    "factor_prune": "因子精简",
    "ensemble_offsets": "多偏移集成",
    "enable_enhanced_features": "增强因子",
    "enable_north_features": "北向资金因子",
    "enable_lhb_features": "龙虎榜因子",
    "enable_consensus_features": "一致预期因子",
    "enable_cashflow_quality_features": "现金流质量因子",
    "enable_consensus_revision_features": "一致预期修正因子",
    "oos_backtest": "OOS回测",
    "oos_backtest_months": "OOS回测月数",
    "bt_top_n": "回测TopN",
    # 盈亏动态持仓
    # ATR 动态阈值与仓位缩放
    # 亏损提前换出二次确认
    # 仓位管理
    "position_sizing": "仓位模式",
    "kelly_vol_window": "Kelly波动窗口",
    "kelly_max_leverage": "Kelly仓位上限",
    "stagger_tranches": "分批调仓批数",
    # 整体持仓止盈
    "enable_early_rebalance_on_empty": "空仓提前调仓",
    "skip_training": "跳过训练",
    "start_model_version": "起始模型版本",
    "no_deploy_train": "禁用部署训练",
}


BATCH_EXPERIMENT_CORE_COLS = [
    "运行ID",
    "最新运行时间",
    "批次ID",
    "批次时间段",
    "最终日期",
    "综合得分",
    "选股综合得分",
    "全周期CAGR",
    "全周期链式最大回撤",
    "全周期链式夏普",
    "跨切分IR",
    "回测胜率",
    "RankIC均值",
    "ICIR",
    "Top30超额均值",
    "Top30最差中位收益",
    "分层单调性(近似)",
    "验证_OOS_IR差距",
    "惩罚覆盖率",
    "替换收益贡献",
    "重点Top20命中率均值",
    "重点Top20收益中位数均值",
    "重点Top30命中率均值",
    "重点Top30收益中位数均值",
]


BATCH_EXPERIMENT_PARAM_CANDIDATES = [
    "标签列",
    "任务类型",
    "滚动频率",
    "训练窗口年数",
    "测试窗口月数",
    "树数量",
    "最大深度",
    "学习率",
    "rank权重TopK",
    "rank权重值",
    "多种子bagging种子",
    "多种子保留比例",
    "多种子最少保留模型数",
    "最佳迭代均值",
    "回测TopN",
    "回测调仓频率",
    "回测卖出时机",
    "市场择时",
    "盈亏动态持仓",
]

# ---------------------------------------------------------------------------
# 训练参数列（来自 write_walk_forward_summary 写入的列名，取每组第一行即可）
# ---------------------------------------------------------------------------
PARAM_COLS = [
    "wf_run_id",
    "batch_run_id",
    "batch_period_label",
    "algorithm",
    "split_count",
    "final_date",
    "wf_start_date",
    "wf_end_date",
    "step",
    "train_window_years",
    "test_window_months",
    "val_ratio",
    "label_column",
    "task",
    "label_transform",
    "n_estimators",
    "max_depth",
    "num_leaves",
    "learning_rate",
    "subsample",
    "colsample_bytree",
    "min_child_weight",
    "gamma",
    "reg_alpha",
    "reg_lambda",
    "early_stopping_rounds",
    "early_stopping_metric",
    "adaptive_best_iter_retrain",
    "adaptive_low_iter_max_retries",
    "ensemble_seeds",
    "ensemble_seed_keep_top_ratio",
    "ensemble_seed_keep_min_models",
    "rank_weight_enabled",
    "rank_weight_topk",
    "rank_weight",
    "time_decay_half_life",
    "freshness_strategy",
    "event_freshness_half_life_days",
    "objective",
    "enable_fundamental",
    "enable_alt",
    "enable_margin",
    "enable_cyq",
    "enable_fund",
    "enable_express",
    "feature_stability_filter",
    "factor_prune",
    "ensemble_offsets",
    "enable_enhanced_features",
    "enable_north_features",
    "enable_lhb_features",
    "enable_consensus_features",
    "enable_cashflow_quality_features",
    "enable_consensus_revision_features",
    "oos_backtest",
    "oos_backtest_months",
    "bt_top_n",
    "bt_rebalance_freq",
    "bt_initial_capital",
    "bt_sell_timing",
    "bt_exclude_st",
    "bt_min_list_days",
    "bt_max_weight_per_stock",
    "bt_max_per_industry",
    "bt_stop_loss_enabled",
    "bt_stop_loss_drawdown_pct",
    "bt_stop_loss_trailing_enabled",
    "bt_stop_loss_trailing_pct",
    "bt_stop_loss_consecutive_limit_down",
    "bt_equity_curve_enabled",
    "bt_equity_curve_drawdown_thresholds",
    "bt_equity_curve_exposure_levels",
    "bt_equity_curve_ma_short",
    "bt_equity_curve_ma_long",
    "bt_equity_curve_recovery_mode",
    "bt_equity_curve_recovery_step",
    "bt_equity_curve_recovery_delay_periods",
    # 行业轮动加权
    "industry_momentum_filter",
    "industry_momentum_bottom_pct",
    "industry_rotation_enhanced",
    "industry_rotation_alpha",
    # 仓位管理
    "position_sizing",
    "kelly_vol_window",
    "kelly_max_leverage",
    "market_regime",
    "market_regime_bear_threshold",
    "market_regime_bear_exposure",
    "market_regime_mode",
    "market_regime_vol_target",
    "market_regime_trend_threshold",
    "market_regime_min_exposure",
    "market_regime_combine_method",
    "market_regime_trend_guard",
    "market_regime_drawdown_guard",
    "market_regime_drawdown_threshold",
    "market_regime_ma250_hard_stop",
    "market_regime_ma250_threshold",
    "market_regime_ma250_exposure",
    "market_regime_ma250_atr_scaling",
    "stagger_tranches",
    "enable_early_rebalance_on_empty",
    "skip_training",
    "start_model_version",
    "no_deploy_train",
]


MODEL_PARAM_KEYS = [
    "algorithm",
    "split_count",
    "final_date",
    "wf_start_date",
    "wf_end_date",
    "step",
    "train_window_years",
    "test_window_months",
    "val_ratio",
    "label_column",
    "task",
    "label_transform",
    "n_estimators",
    "max_depth",
    "num_leaves",
    "learning_rate",
    "subsample",
    "colsample_bytree",
    "min_child_weight",
    "gamma",
    "reg_alpha",
    "reg_lambda",
    "early_stopping_rounds",
    "early_stopping_metric",
    "adaptive_best_iter_retrain",
    "adaptive_low_iter_max_retries",
    "ensemble_seeds",
    "ensemble_seed_keep_top_ratio",
    "ensemble_seed_keep_min_models",
    "rank_weight_enabled",
    "rank_weight_topk",
    "rank_weight",
    "rank_weight_topk_weight_mode",
    "time_decay_half_life",
    "freshness_strategy",
    "event_freshness_half_life_days",
    "objective",
    "enable_fundamental",
    "enable_alt",
    "enable_margin",
    "enable_cyq",
    "enable_fund",
    "enable_express",
    "feature_stability_filter",
    "factor_prune",
    "ensemble_offsets",
    "enable_enhanced_features",
    "enable_north_features",
    "enable_lhb_features",
    "enable_consensus_features",
    "enable_cashflow_quality_features",
    "enable_consensus_revision_features",
    "skip_training",
    "start_model_version",
    "no_deploy_train",
]


SEED_STABILITY_EXCLUDED_MODEL_KEYS = [
    "ensemble_seeds",
    "ensemble_seed_keep_top_ratio",
    "ensemble_seed_keep_min_models",
]


BATCH_EXPERIMENT_EXCLUDED_PARAM_COLS = {
    COL_NAMES["ensemble_seeds"],
    COL_NAMES["ensemble_seed_keep_top_ratio"],
    COL_NAMES["ensemble_seed_keep_min_models"],
}


TRADE_PARAM_KEYS = [
    "oos_backtest",
    "oos_backtest_months",
    "bt_top_n",
    "bt_rebalance_freq",
    "bt_initial_capital",
    "bt_sell_timing",
    "bt_exclude_st",
    "bt_min_list_days",
    "bt_max_weight_per_stock",
    "bt_max_per_industry",
    "bt_stop_loss_enabled",
    "bt_stop_loss_drawdown_pct",
    "bt_stop_loss_trailing_enabled",
    "bt_stop_loss_trailing_pct",
    "bt_stop_loss_consecutive_limit_down",
    "bt_equity_curve_enabled",
    "bt_equity_curve_drawdown_thresholds",
    "bt_equity_curve_exposure_levels",
    "bt_equity_curve_ma_short",
    "bt_equity_curve_ma_long",
    "bt_equity_curve_recovery_mode",
    "bt_equity_curve_recovery_step",
    "bt_equity_curve_recovery_delay_periods",
    "industry_momentum_filter",
    "industry_momentum_bottom_pct",
    "industry_rotation_enhanced",
    "industry_rotation_alpha",
    "position_sizing",
    "kelly_vol_window",
    "kelly_max_leverage",
    "market_regime",
    "market_regime_bear_threshold",
    "market_regime_bear_exposure",
    "market_regime_mode",
    "market_regime_vol_target",
    "market_regime_trend_threshold",
    "market_regime_min_exposure",
    "market_regime_combine_method",
    "market_regime_trend_guard",
    "market_regime_drawdown_guard",
    "market_regime_drawdown_threshold",
    "market_regime_ma250_hard_stop",
    "market_regime_ma250_threshold",
    "market_regime_ma250_exposure",
    "market_regime_ma250_atr_scaling",
    "stagger_tranches",
    "enable_early_rebalance_on_empty",
]


PAIR_CONTEXT_KEYS = ["batch_period_label", "final_date", "split_count"]


MODEL_ALPHA_SCORE_CONFIG = [
    ("选股综合得分均值", 0.30, "high"),
    ("ICIR均值", 0.20, "high"),
    ("Top30超额均值", 0.15, "high"),
    ("Top30最差中位收益", 0.15, "high"),
    ("分层单调性均值", 0.10, "high"),
    ("验证_OOS_IR差距", 0.10, "low"),
]


TRADE_YIELD_SCORE_CONFIG = [
    ("CAGR配对百分位均值", 0.40, "high"),
    ("总收益配对百分位均值", 0.25, "high"),
    ("Calmar配对百分位均值", 0.15, "high"),
    ("夏普配对百分位均值", 0.10, "high"),
    ("胜率配对百分位均值", 0.10, "high"),
]


TRADE_ROBUST_SCORE_CONFIG = [
    ("最大回撤配对百分位均值", 0.35, "high"),
    ("Calmar配对百分位均值", 0.25, "high"),
    ("夏普配对百分位均值", 0.20, "high"),
    ("胜率配对百分位均值", 0.10, "high"),
    ("CAGR最差配对百分位", 0.10, "high"),
]


CANDIDATE_SCORE_CONFIG = [
    ("模型Alpha分", 0.45, "high"),
    ("交易收益分", 0.30, "high"),
    ("交易稳健分", 0.15, "high"),
    ("最差场景防守分", 0.10, "high"),
]


CANDIDATE_MIN_MODEL_ALPHA = 60.0
CANDIDATE_MIN_EFFECTIVE_PAIR_CONTEXTS = 2
CANDIDATE_MIN_CHAIN_MAX_DRAWDOWN = -0.35
CANDIDATE_MIN_CHAIN_CAGR_WORST = -0.05


def _is_missing_param_value(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip().lower() in ("", "nan", "none", "null")
    try:
        return bool(pd.isna(value))
    except Exception:
        return False


def _is_true_param_value(value) -> bool:
    if _is_missing_param_value(value):
        return False
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in ("true", "1", "yes", "y"):
            return True
        if normalized in ("false", "0", "no", "n"):
            return False
    return bool(value)


def _normalize_param_text(value) -> Optional[str]:
    if _is_missing_param_value(value):
        return None
    return str(value).strip()


def _parse_optional_int(value) -> Optional[int]:
    if _is_missing_param_value(value):
        return None
    text = str(value).strip()
    try:
        return int(text)
    except (TypeError, ValueError):
        try:
            return int(float(text))
        except (TypeError, ValueError):
            return None


def _get_open_trade_dates_and_index(
    trade_cal: pd.DataFrame,
) -> tuple[list[str], dict[str, int]]:
    cache_key = id(trade_cal)
    cached = _TRADE_CAL_OPEN_DATES_CACHE.get(cache_key)
    if cached is not None:
        return cached

    open_dates = (
        trade_cal.loc[trade_cal["is_open"] == 1, "cal_date"].astype(str).sort_values().tolist()
    )
    index_map = {trade_date: idx for idx, trade_date in enumerate(open_dates)}
    cached = (open_dates, index_map)
    _TRADE_CAL_OPEN_DATES_CACHE[cache_key] = cached
    return cached


def _find_nearest_trade_date_backward(
    all_trade_dates: list[str], target_date: str
) -> Optional[str]:
    if not all_trade_dates:
        return None
    pos = bisect_right(all_trade_dates, target_date)
    if pos <= 0:
        return None
    return all_trade_dates[pos - 1]


def _collect_rebalance_freq_candidates(
    ordered: pd.DataFrame,
    all_trade_dates: list[str],
    trade_date_index: dict[str, int],
    test_window_months: int,
) -> set[int]:
    probe_indices = ordered["__split_index_int"].astype(int).tolist()
    probe_indices = sorted(set(probe_indices[:3] + probe_indices[-3:]))

    candidates: Optional[set[int]] = None
    for idx in probe_indices:
        row = ordered.loc[ordered["__split_index_int"] == idx].iloc[0]
        test_start = _normalize_param_text(row.get("test_start"))
        test_end = _normalize_param_text(row.get("test_end"))
        if not test_start or not test_end:
            continue

        test_start_idx = trade_date_index.get(test_start)
        test_end_idx = trade_date_index.get(test_end)
        if test_start_idx is None or test_end_idx is None or test_end_idx < test_start_idx:
            continue

        nominal_target = (
            datetime.strptime(test_start, "%Y%m%d") + relativedelta(months=test_window_months)
        ).strftime("%Y%m%d")
        nominal_end = _find_nearest_trade_date_backward(all_trade_dates, nominal_target)
        if nominal_end is None:
            continue

        nominal_end_idx = trade_date_index.get(nominal_end)
        if nominal_end_idx is None or nominal_end_idx < test_start_idx:
            continue

        # 末段可能被 final_date 截断；此时无法由边界稳定反推调仓频率，跳过该 probe。
        if test_end_idx < nominal_end_idx:
            continue

        actual_span = test_end_idx - test_start_idx + 1
        nominal_span = nominal_end_idx - test_start_idx + 1
        min_valid_freq = actual_span - nominal_span + 1

        row_candidates: set[int] = set()
        upper = min(actual_span, _BT_REBALANCE_FREQ_MAX)
        divisor = 1
        while divisor * divisor <= actual_span:
            if actual_span % divisor == 0:
                pair = actual_span // divisor
                if divisor >= min_valid_freq and divisor <= upper:
                    row_candidates.add(divisor)
                if pair >= min_valid_freq and pair <= upper:
                    row_candidates.add(pair)
            divisor += 1

        if not row_candidates:
            continue

        candidates = row_candidates if candidates is None else candidates & row_candidates
        if not candidates:
            break

    return candidates or set()


def _load_trade_cal_for_compare(data_root: Optional[Path]) -> Optional[pd.DataFrame]:
    if data_root is None:
        return None

    cache_key = str(Path(data_root).resolve())
    if cache_key in _TRADE_CAL_CACHE:
        return _TRADE_CAL_CACHE[cache_key]

    try:
        storage = Storage(root_path=str(data_root))
        loader = DataLoader(storage)
        trade_cal = loader.load_clean_trade_cal()
        if trade_cal is None or len(trade_cal) == 0:
            trade_cal = loader.load_trade_cal()
    except Exception as exc:
        logger.warning(f"加载交易日历失败，无法回推 bt_rebalance_freq: {exc}")
        _TRADE_CAL_CACHE[cache_key] = None
        return None

    if trade_cal is None or len(trade_cal) == 0:
        logger.warning("交易日历为空，无法回推 bt_rebalance_freq")
        _TRADE_CAL_CACHE[cache_key] = None
        return None

    _TRADE_CAL_CACHE[cache_key] = trade_cal
    return trade_cal


def _infer_bt_rebalance_freq_from_group(
    group: pd.DataFrame,
    trade_cal: Optional[pd.DataFrame],
) -> Optional[int]:
    if group.empty:
        return None

    if "bt_rebalance_freq" in group.columns:
        for value in group["bt_rebalance_freq"].tolist():
            parsed = _parse_optional_int(value)
            if parsed is not None:
                return parsed

    if trade_cal is None:
        return None

    first = group.iloc[0]
    all_trade_dates, trade_date_index = _get_open_trade_dates_and_index(trade_cal)
    if not all_trade_dates:
        return None

    test_window_months = _parse_optional_int(first.get("test_window_months"))
    if test_window_months is None:
        return None

    ordered = group.copy()
    ordered["__split_index_int"] = pd.to_numeric(ordered.get("split_index"), errors="coerce")
    ordered = ordered.dropna(subset=["__split_index_int"]).sort_values("__split_index_int")

    if ordered.empty:
        return None

    candidates = _collect_rebalance_freq_candidates(
        ordered,
        all_trade_dates,
        trade_date_index,
        test_window_months,
    )
    if not candidates:
        return None

    # 同一批边界约束下，满足条件的最小频率就是实际调仓频率；更大的倍数只是在个别窗口上“碰巧也对齐”。
    return min(candidates)


def _fill_missing_bt_rebalance_freq(
    all_df: pd.DataFrame,
    data_root: Optional[Path],
) -> pd.DataFrame:
    if all_df.empty or "wf_run_id" not in all_df.columns:
        return all_df

    filled = all_df.copy()
    if "bt_rebalance_freq" not in filled.columns:
        filled["bt_rebalance_freq"] = pd.NA

    trade_cal: Optional[pd.DataFrame] = None
    for wf_run_id, group in filled.groupby("wf_run_id", sort=False):
        inferred = _infer_bt_rebalance_freq_from_group(group, trade_cal)
        if inferred is None:
            if trade_cal is None:
                trade_cal = _load_trade_cal_for_compare(data_root)
            inferred = _infer_bt_rebalance_freq_from_group(group, trade_cal)

        if inferred is not None:
            filled.loc[group.index, "bt_rebalance_freq"] = inferred
        else:
            logger.debug(f"无法从 summary 回推 bt_rebalance_freq: {wf_run_id}")

    return filled


def _sanitize_summary_train_params(raw_params: dict) -> dict:
    """兼容历史 summary：按参数是否实际生效清空旧默认值。"""
    params = dict(raw_params)

    def clear(*keys: str) -> None:
        for key in keys:
            if key in params:
                params[key] = None

    if not _is_true_param_value(params.get("oos_backtest")):
        clear(
            "oos_backtest_months",
            "bt_top_n",
            "bt_rebalance_freq",
            "bt_sell_timing",
            "bt_exclude_st",
            "bt_min_list_days",
            "bt_max_weight_per_stock",
            "bt_max_per_industry",
            "bt_stop_loss_enabled",
            "bt_stop_loss_drawdown_pct",
            "bt_stop_loss_trailing_enabled",
            "bt_stop_loss_trailing_pct",
            "bt_stop_loss_consecutive_limit_down",
            "bt_equity_curve_enabled",
            "bt_equity_curve_drawdown_thresholds",
            "bt_equity_curve_exposure_levels",
            "bt_equity_curve_ma_short",
            "bt_equity_curve_ma_long",
            "bt_equity_curve_recovery_mode",
            "bt_equity_curve_recovery_step",
            "bt_equity_curve_recovery_delay_periods",
            "industry_momentum_filter",
            "industry_momentum_bottom_pct",
            "industry_rotation_enhanced",
            "industry_rotation_alpha",
            "position_sizing",
            "kelly_vol_window",
            "kelly_max_leverage",
            "market_regime",
            "market_regime_bear_threshold",
            "market_regime_bear_exposure",
            "market_regime_mode",
            "market_regime_vol_target",
            "market_regime_trend_threshold",
            "market_regime_min_exposure",
            "market_regime_combine_method",
            "market_regime_trend_guard",
            "market_regime_drawdown_guard",
            "market_regime_drawdown_threshold",
            "market_regime_ma250_hard_stop",
            "market_regime_ma250_threshold",
            "market_regime_ma250_exposure",
            "market_regime_ma250_atr_scaling",
            "stagger_tranches",
            "enable_early_rebalance_on_empty",
        )
        return params

    if not _is_true_param_value(params.get("bt_stop_loss_enabled")):
        clear(
            "bt_stop_loss_drawdown_pct",
            "bt_stop_loss_trailing_enabled",
            "bt_stop_loss_trailing_pct",
            "bt_stop_loss_consecutive_limit_down",
        )
    elif not _is_true_param_value(params.get("bt_stop_loss_trailing_enabled")):
        clear("bt_stop_loss_trailing_pct")

    if not _is_true_param_value(params.get("bt_equity_curve_enabled")):
        clear(
            "bt_equity_curve_drawdown_thresholds",
            "bt_equity_curve_exposure_levels",
            "bt_equity_curve_ma_short",
            "bt_equity_curve_ma_long",
            "bt_equity_curve_recovery_mode",
            "bt_equity_curve_recovery_step",
            "bt_equity_curve_recovery_delay_periods",
        )

    if not _is_true_param_value(params.get("industry_momentum_filter")):
        clear("industry_momentum_bottom_pct")

    if not _is_true_param_value(params.get("industry_rotation_enhanced")):
        clear("industry_rotation_alpha")

    if _normalize_param_text(params.get("position_sizing")) not in ("kelly", "half_kelly"):
        clear("kelly_vol_window", "kelly_max_leverage")

    if not _is_true_param_value(params.get("market_regime")):
        clear(
            "market_regime_bear_threshold",
            "market_regime_bear_exposure",
            "market_regime_mode",
            "market_regime_vol_target",
            "market_regime_trend_threshold",
            "market_regime_min_exposure",
            "market_regime_combine_method",
            "market_regime_trend_guard",
            "market_regime_drawdown_guard",
            "market_regime_drawdown_threshold",
        )
    else:
        market_regime_mode = _normalize_param_text(params.get("market_regime_mode"))
        if market_regime_mode == "binary":
            clear(
                "market_regime_vol_target",
                "market_regime_trend_threshold",
                "market_regime_min_exposure",
                "market_regime_combine_method",
                "market_regime_trend_guard",
            )
        elif market_regime_mode == "vol_target":
            clear(
                "market_regime_bear_threshold",
                "market_regime_bear_exposure",
                "market_regime_trend_threshold",
                "market_regime_combine_method",
                "market_regime_trend_guard",
            )
        elif market_regime_mode == "trend":
            clear(
                "market_regime_bear_threshold",
                "market_regime_bear_exposure",
                "market_regime_vol_target",
                "market_regime_combine_method",
                "market_regime_trend_guard",
            )
        elif market_regime_mode == "combined":
            clear(
                "market_regime_bear_threshold",
                "market_regime_bear_exposure",
            )

    if not _is_true_param_value(params.get("market_regime_ma250_hard_stop")):
        clear(
            "market_regime_ma250_threshold",
            "market_regime_ma250_exposure",
            "market_regime_ma250_atr_scaling",
        )

    if not _is_true_param_value(params.get("market_regime_drawdown_guard")):
        clear("market_regime_drawdown_threshold")

    return params


def _sanitize_summary_frame(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    sanitized_rows = [_sanitize_summary_train_params(row.to_dict()) for _, row in df.iterrows()]
    return pd.DataFrame(sanitized_rows, columns=df.columns)


def _concat_summary_frames(frames: list[pd.DataFrame]) -> pd.DataFrame:
    """按旧语义拼接 summary，避免 pandas 对全 NA 列发出 FutureWarning。"""
    if not frames:
        return pd.DataFrame()

    ordered_columns: list[str] = []
    for frame in frames:
        for col in frame.columns:
            if col not in ordered_columns:
                ordered_columns.append(col)

    prepared_frames = []
    for frame in frames:
        if frame.empty:
            continue
        # 显式排除单个 frame 内全 NA 列，保持 pandas 旧版 concat 的 dtype 推断语义。
        prepared_frames.append(frame.dropna(axis=1, how="all"))

    if not prepared_frames:
        return pd.DataFrame(columns=ordered_columns)

    all_df = pd.concat(prepared_frames, ignore_index=True)
    return all_df.reindex(columns=ordered_columns)


# ---------------------------------------------------------------------------
# 综合得分配置：(英文列键, 权重, 方向)
#   "high"    → 值越大越好
#   "low"     → 值越小越好
#   "abs_low" → 绝对值越小越好
# 权重之和应为 1.0
# ---------------------------------------------------------------------------
SCORE_CONFIG = [
    # ── 回测指标（60%）：真实组合模拟，最直接反映参数优劣 ──────────
    ("chain_cagr", 0.20, "high"),  # 全周期串联 CAGR：最终最该看的盈利能力
    ("bt_win_rate", 0.15, "high"),  # 回测胜率：各切分正收益占比
    ("bt_sharpe_mean", 0.15, "high"),  # 回测夏普均值：风险收益比
    ("chain_max_drawdown", 0.10, "high"),  # 全周期链式最大回撤：真实风险下限
    # ── 统计指标（32%）：辅助验证，防止回测过拟合 ─────────────────
    ("oos_cross_split_ir", 0.10, "high"),  # 跨切分IR：核心稳健性
    ("oos_top30_win_rate", 0.05, "high"),  # Top30胜率：策略持续性
    ("oos_top30_worst_median", 0.05, "high"),  # Top30最差切分：抗压能力
    ("oos_rankic_ir_trend", 0.05, "high"),  # IC趋势：alpha是否在衰减
    ("oos_top30_median_mean", 0.03, "high"),  # Top30中位收益：核心盈利能力
    ("oos_top30_lift_mean", 0.02, "high"),  # Top30超额：纯选股能力
    ("oos_top30_skew_score_mean", 0.02, "abs_low"),  # 偏斜度：极端日干扰风险（绝对值越小越好）
    # ── 训练质量（8%）：过拟合检测 ────────────────────────────────
    ("train_val_ir_gap", 0.08, "low"),  # 验证_OOS差距：过拟合风险（越小越好）
]


def load_chain_metrics(
    raw_dir: Optional[Path], wf_run_id: str, source_dir: Optional[Path] = None
) -> dict:
    """读取 chain_nav 并计算全周期指标。"""
    empty = {
        "chain_total_return": None,
        "chain_cagr": None,
        "chain_max_drawdown": None,
        "chain_sharpe": None,
        "chain_trading_days": None,
    }
    effective_raw_dir = source_dir or raw_dir
    if effective_raw_dir is None:
        return empty

    chain_path = effective_raw_dir / f"chain_nav_{wf_run_id}.csv"
    if not chain_path.exists():
        return empty

    try:
        chain_df = pd.read_csv(chain_path, encoding="utf-8-sig")
    except Exception as exc:
        logger.warning(f"读取 chain_nav 失败: {chain_path.name} — {exc}")
        return empty

    if chain_df.empty or "nav" not in chain_df.columns:
        return empty

    nav = pd.to_numeric(chain_df["nav"], errors="coerce").dropna()
    if nav.empty or nav.iloc[0] == 0:
        return empty

    trading_days = len(nav)
    years = trading_days / 252 if trading_days > 0 else 0
    total_return = nav.iloc[-1] / nav.iloc[0] - 1
    # 简单年化收益率（不假设收益再投入）
    cagr = (total_return / years) if years > 0 else None
    cummax = nav.cummax()
    drawdown = (nav - cummax) / cummax
    max_drawdown = drawdown.min() if not drawdown.empty else None
    daily_ret = nav.pct_change().dropna()
    volatility = daily_ret.std() * (252**0.5) if len(daily_ret) > 1 else None
    sharpe = None
    if cagr is not None and volatility is not None and volatility > 0:
        sharpe = (cagr - 0.03) / volatility

    return {
        "chain_total_return": round(total_return, 6),
        "chain_cagr": round(cagr, 6) if cagr is not None else None,
        "chain_max_drawdown": round(max_drawdown, 6) if max_drawdown is not None else None,
        "chain_sharpe": round(sharpe, 4) if sharpe is not None else None,
        "chain_trading_days": trading_days,
    }


def load_all_summaries_from_raw_dirs(
    raw_dirs: list[Path],
    data_root: Optional[Path] = None,
) -> pd.DataFrame:
    """从一个或多个 raw 目录加载 walk_forward 汇总 CSV。"""
    existing_raw_dirs = [Path(raw_dir) for raw_dir in raw_dirs if Path(raw_dir).exists()]
    if len(existing_raw_dirs) == 0:
        logger.warning("未找到任何可用汇总目录")
        return pd.DataFrame()

    csv_files: list[tuple[Path, Path]] = []
    for raw_dir in existing_raw_dirs:
        csv_files.extend(
            (raw_dir, csv_file) for csv_file in sorted(raw_dir.glob("walk_forward_summary_*.csv"))
        )

    if len(csv_files) == 0:
        joined_dirs = ", ".join(str(raw_dir) for raw_dir in existing_raw_dirs)
        logger.warning(f"未找到任何汇总CSV: {joined_dirs}")
        return pd.DataFrame()

    logger.info(f"找到 {len(csv_files)} 个汇总CSV文件")
    frames = []
    for raw_dir, f in csv_files:
        try:
            df = pd.read_csv(
                f,
                encoding="utf-8-sig",
                dtype=SUMMARY_CSV_DTYPE,
            )
            df = _sanitize_summary_frame(df)
            df["_source_file"] = f.name
            df["_source_dir"] = str(raw_dir)
            frames.append(df)
            logger.debug(f"  已加载: {f.name}（{len(df)} 行）")
        except Exception as e:
            logger.warning(f"  跳过（读取失败）: {f.name} — {e}")

    if not frames:
        return pd.DataFrame()

    all_df = _concat_summary_frames(frames)
    all_df = _fill_missing_bt_rebalance_freq(all_df, data_root)
    logger.info(
        f"合并后总行数: {len(all_df)}，unique wf_run_id: {all_df['wf_run_id'].nunique() if 'wf_run_id' in all_df.columns else '?'}"
    )
    return all_df


def load_all_summaries(raw_dir: Path, data_root: Optional[Path] = None) -> pd.DataFrame:
    """兼容旧调用：从单个 raw 目录加载汇总CSV。"""
    return load_all_summaries_from_raw_dirs([raw_dir], data_root=data_root)


def build_auto_compare_jobs(data_root: Path) -> list[dict]:
    """构建无参模式下的自动扫描任务。"""
    walk_forward_root = data_root / "walk_forward"
    raw_dir = walk_forward_root / "raw"
    batches_root = walk_forward_root / "batches"
    batch_raw_dirs = (
        sorted(path for path in batches_root.glob("*/raw") if path.is_dir())
        if batches_root.exists()
        else []
    )

    jobs = [
        {
            "label": "raw",
            "raw_dirs": [raw_dir],
            "output_path": walk_forward_root / "wf_comparison_raw.xlsx",
        }
    ]
    if batch_raw_dirs:
        jobs.append(
            {
                "label": "batches",
                "raw_dirs": batch_raw_dirs,
                "output_path": walk_forward_root / "wf_comparison_batches.xlsx",
            }
        )
    return jobs


def aggregate_run(group: pd.DataFrame) -> dict:
    """对单个 wf_run_id 的所有 split 行进行聚合，返回一行对比指标"""
    row = {}
    n = len(group)
    row["n_splits"] = n

    # 模型版本范围（min~max）
    if "model_version" in group.columns:
        mv = group["model_version"].dropna()
        if len(mv):
            row["model_version_range"] = f"{int(mv.min())}~{int(mv.max())}"
        else:
            row["model_version_range"] = None
    else:
        row["model_version_range"] = None

    # -----------------------------------------------------------------------
    # OOS 性能指标（来自 test_daily_metrics 展开列）
    # -----------------------------------------------------------------------
    def safe_mean(col):
        return group[col].mean() if col in group.columns else None

    def safe_std(col):
        return group[col].std() if col in group.columns else None

    def safe_min(col):
        return group[col].min() if col in group.columns else None

    def safe_max(col):
        return group[col].max() if col in group.columns else None

    # KEY 重点字段（前置展示）
    row["KEY_说明"] = "重点: hit rate=TopK逐日平均收益>0占比; list=最新OOS日期预测名单"
    row["KEY_Top20_list"] = None
    row["KEY_Top30_list"] = None
    if "split_index" in group.columns:
        sorted_group = group.copy()
        sorted_group["__split_index_int"] = pd.to_numeric(
            sorted_group["split_index"], errors="coerce"
        )
        sorted_group = sorted_group.sort_values("__split_index_int")
    else:
        sorted_group = group.copy()
    if "KEY_Top20_list" in sorted_group.columns:
        top20_list = sorted_group["KEY_Top20_list"].dropna()
        row["KEY_Top20_list"] = str(top20_list.iloc[-1]) if len(top20_list) else None
    if "KEY_Top30_list" in sorted_group.columns:
        top30_list = sorted_group["KEY_Top30_list"].dropna()
        row["KEY_Top30_list"] = str(top30_list.iloc[-1]) if len(top30_list) else None

    key20_hit = safe_mean("KEY_Top20_hit_rate")
    key20_med = safe_mean("KEY_Top20_avg_return_median")
    key30_hit = safe_mean("KEY_Top30_hit_rate")
    key30_med = safe_mean("KEY_Top30_avg_return_median")
    row["key_top20_hit_rate_mean"] = round(key20_hit, 4) if key20_hit is not None else None
    row["key_top20_avg_return_median_mean"] = round(key20_med, 6) if key20_med is not None else None
    row["key_top30_hit_rate_mean"] = round(key30_hit, 4) if key30_hit is not None else None
    row["key_top30_avg_return_median_mean"] = round(key30_med, 6) if key30_med is not None else None

    # OOS RankIC IR
    oos_ir_series = (
        group["daily_rankic_ir"] if "daily_rankic_ir" in group.columns else pd.Series(dtype=float)
    )
    oos_ir_mean = oos_ir_series.mean() if len(oos_ir_series) else None
    oos_ir_std = oos_ir_series.std() if len(oos_ir_series) > 1 else None
    row["oos_rankic_ir_mean"] = round(oos_ir_mean, 4) if oos_ir_mean is not None else None
    row["oos_rankic_ir_std"] = round(oos_ir_std, 4) if oos_ir_std is not None else None
    row["oos_cross_split_ir"] = (
        round(oos_ir_mean / oos_ir_std, 3)
        if (oos_ir_mean and oos_ir_std and oos_ir_std != 0)
        else None
    )

    # RankIC 均值与 ICIR（纯选股能力核心指标）
    rankic_mean_series = (
        group["daily_rankic_mean"].dropna()
        if "daily_rankic_mean" in group.columns
        else pd.Series(dtype=float)
    )
    rankic_std_series = (
        group["daily_rankic_std"].dropna()
        if "daily_rankic_std" in group.columns
        else pd.Series(dtype=float)
    )
    rankic_mean = rankic_mean_series.mean() if len(rankic_mean_series) else None
    rankic_std = rankic_std_series.mean() if len(rankic_std_series) else None
    row["daily_rankic_mean"] = round(rankic_mean, 6) if rankic_mean is not None else None
    row["icir"] = (
        round(rankic_mean / rankic_std, 4)
        if (rankic_mean is not None and rankic_std is not None and rankic_std != 0)
        else None
    )

    # OOS RankIC 衰减检测（最近3个split均值 - 最早3个split均值）
    if len(oos_ir_series) >= 6:
        sorted_ir = (
            group.sort_values("split_index")["daily_rankic_ir"]
            if "daily_rankic_ir" in group.columns
            else oos_ir_series
        )
        row["oos_rankic_ir_trend"] = round(
            sorted_ir.iloc[-3:].mean() - sorted_ir.iloc[:3].mean(), 4
        )
    else:
        row["oos_rankic_ir_trend"] = None

    # Top30 指标（以中位数为核心，不受极端日干扰）
    med30_col = "diagnostic_Top30_逐日均值_50分位"
    mean30_col = "diagnostic_Top30_逐日均值的均值"
    std30_col = "diagnostic_Top30_逐日均值的标准差"
    lift30_col = "diagnostic_Top30_相对全市场提升_均值"

    if med30_col in group.columns:
        med30_series = group[med30_col].dropna()
        row["oos_top30_median_mean"] = round(med30_series.mean(), 6) if len(med30_series) else None
        row["oos_top30_win_rate"] = (
            round((med30_series > 0).mean(), 3) if len(med30_series) else None
        )
        row["oos_top30_worst_median"] = round(med30_series.min(), 6) if len(med30_series) else None
    else:
        row["oos_top30_median_mean"] = row["oos_top30_win_rate"] = row["oos_top30_worst_median"] = (
            None
        )

    # Top30 偏斜度（均值/中位数 gap，衡量是否被极端日驱动）
    if all(c in group.columns for c in [mean30_col, med30_col, std30_col]):
        valid = group[[mean30_col, med30_col, std30_col]].dropna()
        if len(valid):
            skew_scores = (valid[mean30_col] - valid[med30_col]) / valid[std30_col].replace(
                0, np.nan
            )
            row["oos_top30_skew_score_mean"] = round(skew_scores.mean(), 3)
        else:
            row["oos_top30_skew_score_mean"] = None
    else:
        row["oos_top30_skew_score_mean"] = None

    row["oos_top30_lift_mean"] = (
        round(safe_mean(lift30_col), 6) if safe_mean(lift30_col) is not None else None
    )

    # Top100 指标
    med100_col = "diagnostic_Top100_逐日均值_50分位"
    if med100_col in group.columns:
        med100_series = group[med100_col].dropna()
        row["oos_top100_median_mean"] = (
            round(med100_series.mean(), 6) if len(med100_series) else None
        )
        row["oos_top100_win_rate"] = (
            round((med100_series > 0).mean(), 3) if len(med100_series) else None
        )
    else:
        row["oos_top100_median_mean"] = row["oos_top100_win_rate"] = None

    # Top300 指标
    med300_col = "diagnostic_Top300_逐日均值_50分位"
    if med300_col in group.columns:
        med300_series = group[med300_col].dropna()
        row["oos_top300_median_mean"] = (
            round(med300_series.mean(), 6) if len(med300_series) else None
        )
        row["oos_top300_win_rate"] = (
            round((med300_series > 0).mean(), 3) if len(med300_series) else None
        )
    else:
        row["oos_top300_median_mean"] = row["oos_top300_win_rate"] = None

    # 分层单调性近似评分（Top30/100/300 中位收益应随覆盖范围扩大而递减）
    monotonic_inputs = [
        (30, row.get("oos_top30_median_mean")),
        (100, row.get("oos_top100_median_mean")),
        (300, row.get("oos_top300_median_mean")),
    ]
    monotonic_inputs = [(k, v) for k, v in monotonic_inputs if v is not None and pd.notna(v)]
    if len(monotonic_inputs) >= 2:
        bucket_sizes = np.array([k for k, _ in monotonic_inputs], dtype=float)
        bucket_returns = np.array([v for _, v in monotonic_inputs], dtype=float)
        if np.allclose(bucket_returns, bucket_returns[0]):
            row["selection_monotonicity"] = 0.5
        else:
            corr = np.corrcoef(bucket_sizes, bucket_returns)[0, 1]
            if pd.notna(corr):
                row["selection_monotonicity"] = round(float(np.clip((1 - corr) / 2, 0.0, 1.0)), 4)
            else:
                row["selection_monotonicity"] = None
    else:
        row["selection_monotonicity"] = None

    # -----------------------------------------------------------------------
    # OOS 回测指标（来自 run_oos_backtest 写入的 bt_* 列）
    # -----------------------------------------------------------------------
    if "bt_total_return" in group.columns:
        bt_ret = group["bt_total_return"].dropna()
        bt_ar = (
            group["bt_annual_return"].dropna()
            if "bt_annual_return" in group.columns
            else pd.Series(dtype=float)
        )
        bt_sh = (
            group["bt_sharpe"].dropna() if "bt_sharpe" in group.columns else pd.Series(dtype=float)
        )
        bt_md = (
            group["bt_max_drawdown"].dropna()
            if "bt_max_drawdown" in group.columns
            else pd.Series(dtype=float)
        )
        bt_cal = (
            group["bt_calmar"].dropna() if "bt_calmar" in group.columns else pd.Series(dtype=float)
        )
        bt_vol = (
            group["bt_volatility"].dropna()
            if "bt_volatility" in group.columns
            else pd.Series(dtype=float)
        )

        row["bt_total_return_mean"] = round(bt_ret.mean(), 6) if len(bt_ret) else None
        row["bt_annual_return_mean"] = round(bt_ar.mean(), 6) if len(bt_ar) else None
        row["bt_sharpe_mean"] = round(bt_sh.mean(), 4) if len(bt_sh) else None
        row["bt_max_drawdown_worst"] = round(bt_md.min(), 6) if len(bt_md) else None
        row["bt_calmar_mean"] = round(bt_cal.mean(), 4) if len(bt_cal) else None
        row["bt_volatility_mean"] = round(bt_vol.mean(), 6) if len(bt_vol) else None
        row["bt_win_rate"] = round((bt_ret > 0).mean(), 3) if len(bt_ret) else None
        bt_gate_block = (
            group["bt_signal_confidence_block_rate"].dropna()
            if "bt_signal_confidence_block_rate" in group.columns
            else pd.Series(dtype=float)
        )
        bt_gate_exposure = (
            group["bt_signal_confidence_avg_exposure"].dropna()
            if "bt_signal_confidence_avg_exposure" in group.columns
            else pd.Series(dtype=float)
        )
        bt_gate_score = (
            group["bt_signal_confidence_avg_score"].dropna()
            if "bt_signal_confidence_avg_score" in group.columns
            else pd.Series(dtype=float)
        )
        row["bt_signal_confidence_block_rate_mean"] = (
            round(bt_gate_block.mean(), 6) if len(bt_gate_block) else None
        )
        row["bt_signal_confidence_avg_exposure_mean"] = (
            round(bt_gate_exposure.mean(), 6) if len(bt_gate_exposure) else None
        )
        row["bt_signal_confidence_avg_score_mean"] = (
            round(bt_gate_score.mean(), 6) if len(bt_gate_score) else None
        )
    else:
        for k in [
            "bt_total_return_mean",
            "bt_annual_return_mean",
            "bt_sharpe_mean",
            "bt_max_drawdown_worst",
            "bt_calmar_mean",
            "bt_volatility_mean",
            "bt_win_rate",
            "bt_signal_confidence_block_rate_mean",
            "bt_signal_confidence_avg_exposure_mean",
            "bt_signal_confidence_avg_score_mean",
        ]:
            row[k] = None

    # -----------------------------------------------------------------------
    # 训练质量指标
    # -----------------------------------------------------------------------
    # 验证集 RankIC IR（val_rankic_ir 列，每 split 一个值）
    if "val_rankic_ir" in group.columns:
        val_ir_series = group["val_rankic_ir"].dropna()
        row["val_rankic_ir_mean"] = round(val_ir_series.mean(), 4) if len(val_ir_series) else None
        # 泛化差距（val IR 越接近 oos IR 越好；负值说明 oos 反而更好，通常是正常的）
        if row["val_rankic_ir_mean"] is not None and row["oos_rankic_ir_mean"] is not None:
            row["train_val_ir_gap"] = round(
                row["val_rankic_ir_mean"] - row["oos_rankic_ir_mean"], 4
            )
        else:
            row["train_val_ir_gap"] = None
    else:
        row["val_rankic_ir_mean"] = row["train_val_ir_gap"] = None

    # 最佳迭代次数统计
    if "best_iteration" in group.columns:
        bi = group["best_iteration"].dropna()
        row["best_iter_mean"] = round(bi.mean(), 1) if len(bi) else None
        row["best_iter_min"] = int(bi.min()) if len(bi) else None
        row["best_iter_max"] = int(bi.max()) if len(bi) else None
        row["best_iter_std"] = round(bi.std(), 1) if len(bi) > 1 else None
    else:
        row["best_iter_mean"] = row["best_iter_min"] = row["best_iter_max"] = row[
            "best_iter_std"
        ] = None

    return row


def build_comparison_table(all_df: pd.DataFrame, raw_dir: Optional[Path] = None) -> pd.DataFrame:
    """构建对比表（行=run，列=聚合指标+训练参数）"""
    if "wf_run_id" not in all_df.columns:
        logger.error("汇总CSV中缺少 wf_run_id 列，无法分组")
        return pd.DataFrame()

    rows = []
    for wf_run_id, group in all_df.groupby("wf_run_id", sort=False):
        source_dir = None
        if "_source_dir" in group.columns:
            source_values = group["_source_dir"].dropna().astype(str)
            if len(source_values):
                source_dir = Path(source_values.iloc[0])

        # 聚合性能指标
        agg = aggregate_run(group)
        agg.update(load_chain_metrics(raw_dir, wf_run_id, source_dir=source_dir))
        agg["wf_run_id"] = wf_run_id

        # 追加训练参数（取第一行即可，同一 run 内所有 split 参数相同）
        first = group.iloc[0]
        for col in PARAM_COLS:
            if col in first.index and col != "wf_run_id":
                agg[col] = first[col]
            elif col != "wf_run_id":
                agg[col] = None

        rows.append(agg)

    if not rows:
        return pd.DataFrame()

    # 列顺序：wf_run_id → 参与评分的指标（按权重降序）→ 非评分指标 → 训练参数
    scored_cols = [col for col, _w, _d in sorted(SCORE_CONFIG, key=lambda x: -x[1])]

    non_scored_metric_cols = [
        "n_splits",
        "model_version_range",
        # 选股指标组合补充
        "daily_rankic_mean",
        "icir",
        "selection_monotonicity",
        # 全周期串联补充
        "chain_total_return",
        "chain_sharpe",
        "chain_trading_days",
        # 回测补充
        "bt_annual_return_mean",
        "bt_calmar_mean",
        "bt_total_return_mean",
        "bt_max_drawdown_worst",
        "bt_volatility_mean",
        "bt_signal_confidence_block_rate_mean",
        "bt_signal_confidence_avg_exposure_mean",
        "bt_signal_confidence_avg_score_mean",
        # 统计补充
        "oos_rankic_ir_mean",
        "oos_rankic_ir_std",
        "oos_top100_median_mean",
        "oos_top100_win_rate",
        "oos_top300_median_mean",
        "oos_top300_win_rate",
        # 训练质量补充
        "val_rankic_ir_mean",
        "best_iter_mean",
        "best_iter_min",
        "best_iter_max",
        "best_iter_std",
    ]

    param_cols_ordered = [c for c in PARAM_COLS if c != "wf_run_id"]

    key_cols = [
        "key_top20_hit_rate_mean",
        "key_top20_avg_return_median_mean",
        "key_top30_hit_rate_mean",
        "key_top30_avg_return_median_mean",
    ]
    all_cols = (
        [
            "wf_run_id",
            "max_depth",
            "learning_rate",
            "rank_weight_topk",
            "rank_weight",
        ]
        + key_cols
        + scored_cols
        + non_scored_metric_cols
        + param_cols_ordered
    )
    df = pd.DataFrame(rows)
    # 只保留存在的列，并按首次出现去重，避免重复列名触发后续 reindex 异常。
    final_cols = []
    for col in all_cols:
        if col in df.columns and col not in final_cols:
            final_cols.append(col)
    df = df[final_cols]

    df = df.reset_index(drop=True)

    # 列名改为中文
    df = df.rename(columns={k: v for k, v in COL_NAMES.items() if k in df.columns})

    return df


def _build_period_label(row: pd.Series) -> str:
    """优先使用批量脚本传入的时间段标签，否则退回到起止日期。"""
    batch_period_label = row.get(COL_NAMES["batch_period_label"])
    if pd.notna(batch_period_label) and str(batch_period_label).strip():
        return str(batch_period_label)

    wf_start = row.get(COL_NAMES["wf_start_date"])
    wf_end = row.get(COL_NAMES["wf_end_date"])
    if pd.notna(wf_start) and pd.notna(wf_end):
        return f"{wf_start}~{wf_end}"
    return "未标注"


def _extract_run_timestamp(run_id: str) -> str:
    """从 wf_run_id 中提取时间戳（YYYYMMDDHHMMSS），用于判定最新 run。"""
    parts = str(run_id).strip().split("_")
    if len(parts) < 3:
        return ""
    date_part, time_part = parts[1], parts[2]
    if len(date_part) == 8 and len(time_part) == 6 and date_part.isdigit() and time_part.isdigit():
        return date_part + time_part
    return ""


def _latest_run_timestamp_from_text(value) -> str:
    """从运行ID或运行ID列表中提取最新时间戳。"""
    text = str(value) if value is not None and pd.notna(value) else ""
    max_ts = ""
    for match in re.finditer(r"wf_(\d{8})_(\d{6})", text):
        ts = match.group(1) + match.group(2)
        if ts > max_ts:
            max_ts = ts
    if not max_ts:
        max_ts = _extract_run_timestamp(text)
    return max_ts


def _format_run_timestamp(ts: str) -> str:
    if not ts or len(ts) != 14:
        return ""
    return f"{ts[0:4]}-{ts[4:6]}-{ts[6:8]} {ts[8:10]}:{ts[10:12]}:{ts[12:14]}"


def sort_by_latest_run_time(
    df: pd.DataFrame,
    source_col: str,
    secondary_sort_cols: Optional[list[str]] = None,
) -> pd.DataFrame:
    """按来源列中的最新 wf_run_id 时间倒序排序，并添加可读的最新运行时间列。"""
    if df.empty or source_col not in df.columns:
        return df

    result = df.copy()
    result["__latest_run_ts"] = result[source_col].map(_latest_run_timestamp_from_text)
    result["最新运行时间"] = result["__latest_run_ts"].map(_format_run_timestamp)
    cols = list(result.columns)
    if "最新运行时间" in cols and source_col in cols:
        cols.remove("最新运行时间")
        cols.insert(cols.index(source_col) + 1, "最新运行时间")
        result = result[cols]
    sort_cols = ["__latest_run_ts"]
    ascending = [False]
    if secondary_sort_cols:
        for col in secondary_sort_cols:
            if col in result.columns:
                sort_cols.append(col)
                ascending.append(False)
    result = result.sort_values(sort_cols, ascending=ascending, na_position="last")
    result = result.drop(columns=["__latest_run_ts"])
    return result.reset_index(drop=True)


def build_period_stability_table(comp_df: pd.DataFrame) -> pd.DataFrame:
    """按参数组合跨时间段聚合，输出稳定性汇总。"""
    if comp_df.empty:
        return pd.DataFrame()

    varying_cols = {
        COL_NAMES["wf_run_id"],
        COL_NAMES["wf_start_date"],
        COL_NAMES["wf_end_date"],
        COL_NAMES["batch_period_label"],
        COL_NAMES["split_count"],
        COL_NAMES["final_date"],
        COL_NAMES["start_model_version"],
    }
    metric_cols = {
        "综合得分",
        COL_NAMES["chain_cagr"],
        COL_NAMES["chain_max_drawdown"],
        COL_NAMES["chain_sharpe"],
        COL_NAMES["oos_cross_split_ir"],
        COL_NAMES["bt_win_rate"],
    }
    group_cols = []
    for key in PARAM_COLS:
        if key == "wf_run_id" or key in SEED_STABILITY_EXCLUDED_MODEL_KEYS:
            continue
        col = COL_NAMES.get(key, key)
        if col and col in comp_df.columns and col not in varying_cols and col not in metric_cols:
            group_cols.append(col)

    if not group_cols:
        return pd.DataFrame()

    working_df = comp_df.copy()
    working_df["__时间段标签"] = working_df.apply(_build_period_label, axis=1)
    run_id_col = COL_NAMES["wf_run_id"]
    working_df["__run_ts"] = working_df[run_id_col].map(_extract_run_timestamp)

    # 同一参数组、同一时间段可能会有多次重复 run（例如扫描了未生效参数）。
    # 这里先按 run 时间戳倒序去重，只保留每个时间段最新的一条，避免时间段数被重复放大。
    dedup_subset = group_cols + ["__时间段标签"]
    working_df = working_df.sort_values(
        ["__run_ts", run_id_col],
        ascending=[False, False],
        na_position="last",
    ).drop_duplicates(subset=dedup_subset, keep="first")

    rows = []
    for _, group in working_df.groupby(group_cols, dropna=False, sort=False):
        if len(group) <= 1:
            continue

        ordered_group = group.sort_values(
            ["__时间段标签", COL_NAMES["wf_run_id"]],
            ascending=[True, True],
            na_position="last",
        )

        score_series = pd.to_numeric(group.get("综合得分"), errors="coerce")
        cagr_series = pd.to_numeric(group.get(COL_NAMES["chain_cagr"]), errors="coerce")
        drawdown_series = pd.to_numeric(group.get(COL_NAMES["chain_max_drawdown"]), errors="coerce")
        ir_series = pd.to_numeric(group.get(COL_NAMES["oos_cross_split_ir"]), errors="coerce")
        win_rate_series = pd.to_numeric(group.get(COL_NAMES["bt_win_rate"]), errors="coerce")
        sharpe_series = pd.to_numeric(group.get(COL_NAMES["chain_sharpe"]), errors="coerce")

        score_std = score_series.std()
        cagr_std = cagr_series.std()
        ir_std = ir_series.std()

        score_penalty = 0.0 if pd.isna(score_std) else min(max(score_std / 20.0, 0.0), 1.0)
        cagr_penalty = 0.0 if pd.isna(cagr_std) else min(max(cagr_std / 0.2, 0.0), 1.0)
        ir_penalty = 0.0 if pd.isna(ir_std) else min(max(ir_std / 1.0, 0.0), 1.0)
        stability_score = round(
            (1 - (0.4 * score_penalty + 0.3 * cagr_penalty + 0.3 * ir_penalty)) * 100,
            1,
        )

        row = {col: group.iloc[0][col] for col in group_cols}
        row.update(
            {
                COL_NAMES["period_count"]: len(group),
                COL_NAMES["period_labels"]: " | ".join(
                    ordered_group["__时间段标签"].astype(str).tolist()
                ),
                COL_NAMES["run_id_list"]: " | ".join(
                    f"{period}:{run_id}"
                    for period, run_id in zip(
                        ordered_group["__时间段标签"].astype(str),
                        ordered_group[COL_NAMES["wf_run_id"]].astype(str),
                    )
                ),
                COL_NAMES["score_mean"]: (
                    round(score_series.mean(), 2) if score_series.notna().any() else None
                ),
                COL_NAMES["score_std"]: round(score_std, 2) if pd.notna(score_std) else None,
                COL_NAMES["score_min"]: (
                    round(score_series.min(), 2) if score_series.notna().any() else None
                ),
                COL_NAMES["score_max"]: (
                    round(score_series.max(), 2) if score_series.notna().any() else None
                ),
                COL_NAMES["chain_cagr_mean"]: (
                    round(cagr_series.mean(), 6) if cagr_series.notna().any() else None
                ),
                COL_NAMES["chain_cagr_std"]: round(cagr_std, 6) if pd.notna(cagr_std) else None,
                COL_NAMES["chain_cagr_min"]: (
                    round(cagr_series.min(), 6) if cagr_series.notna().any() else None
                ),
                COL_NAMES["chain_max_drawdown_mean"]: (
                    round(drawdown_series.mean(), 6) if drawdown_series.notna().any() else None
                ),
                COL_NAMES["chain_max_drawdown_worst"]: (
                    round(drawdown_series.min(), 6) if drawdown_series.notna().any() else None
                ),
                COL_NAMES["oos_cross_split_ir_mean"]: (
                    round(ir_series.mean(), 4) if ir_series.notna().any() else None
                ),
                COL_NAMES["oos_cross_split_ir_std"]: round(ir_std, 4) if pd.notna(ir_std) else None,
                COL_NAMES["bt_win_rate_mean"]: (
                    round(win_rate_series.mean(), 4) if win_rate_series.notna().any() else None
                ),
                COL_NAMES["bt_win_rate_min"]: (
                    round(win_rate_series.min(), 4) if win_rate_series.notna().any() else None
                ),
                COL_NAMES["chain_sharpe_mean"]: (
                    round(sharpe_series.mean(), 4) if sharpe_series.notna().any() else None
                ),
                COL_NAMES["chain_sharpe_std"]: (
                    round(sharpe_series.std(), 4) if sharpe_series.notna().sum() > 1 else None
                ),
                COL_NAMES["stability_score"]: max(stability_score, 0.0),
            }
        )
        rows.append(row)

    if not rows:
        return pd.DataFrame()

    ordered_cols = [
        COL_NAMES["period_count"],
        COL_NAMES["period_labels"],
        COL_NAMES["run_id_list"],
        COL_NAMES["stability_score"],
        COL_NAMES["score_mean"],
        COL_NAMES["score_std"],
        COL_NAMES["score_min"],
        COL_NAMES["score_max"],
        COL_NAMES["chain_cagr_mean"],
        COL_NAMES["chain_cagr_std"],
        COL_NAMES["chain_cagr_min"],
        COL_NAMES["chain_max_drawdown_mean"],
        COL_NAMES["chain_max_drawdown_worst"],
        COL_NAMES["oos_cross_split_ir_mean"],
        COL_NAMES["oos_cross_split_ir_std"],
        COL_NAMES["bt_win_rate_mean"],
        COL_NAMES["bt_win_rate_min"],
        COL_NAMES["chain_sharpe_mean"],
        COL_NAMES["chain_sharpe_std"],
    ]
    ordered_cols += [col for col in group_cols if col not in ordered_cols]
    result = pd.DataFrame(rows)
    result = result[[col for col in ordered_cols if col in result.columns]]
    result = sort_by_latest_run_time(
        result, COL_NAMES["run_id_list"], [COL_NAMES["stability_score"]]
    )
    return result


def compute_composite_score(df: pd.DataFrame) -> pd.Series:
    """计算综合得分（0~100，越高越好）

    方法：对 SCORE_CONFIG 中的每个指标在当前实验集内做百分位排名（0~1），
    按权重加权求和后乘以 100。

    设计原则：
    - 百分位排名（percentile rank）完全回避量纲差异，结果仅反映相对优劣
    - NaN 值视为中性（百分位 0.5），不奖励也不惩罚
    - 单个实验时各指标百分位均为 0.5，得分固定为 50.0
    - ascending=True  时：最大值 → 百分位 1.0（高好）
    - ascending=False 时：最小值 → 百分位 1.0（低好）
    """
    n = len(df)
    weighted_pct = pd.Series(0.0, index=df.index)
    total_weight = 0.0

    for eng_key, weight, direction in SCORE_CONFIG:
        col = COL_NAMES.get(eng_key)
        if col is None or col not in df.columns:
            continue

        s = pd.to_numeric(df[col], errors="coerce")

        if direction == "abs_low":
            s = s.abs()
            ascending = False  # 最大绝对值 → rank 1 → 百分位低 → 得分低 ✓
        elif direction == "low":
            ascending = False  # 最大值 → rank 1 → 百分位低 → 得分低 ✓
        else:  # "high"
            ascending = True  # 最小值 → rank 1 → 百分位低 → 大值得高分 ✓

        if n > 0:
            # rank(ascending=True): 最小→1, 最大→n → pct=rank/n
            pct = s.rank(ascending=ascending, method="average", na_option="keep") / n
            pct = pct.fillna(0.5)
        else:
            pct = pd.Series(0.5, index=df.index)

        weighted_pct += weight * pct
        total_weight += weight

    if total_weight > 0:
        score = (weighted_pct / total_weight) * 100
    else:
        score = pd.Series(50.0, index=df.index)

    return score.round(1)


def compute_selection_score(df: pd.DataFrame) -> pd.Series:
    """计算选股综合得分（0~100，越高越好）。

    指标与权重（选股优先版）：
    - RankIC均值: 30%
    - ICIR: 30%
    - Top30超额均值: 40%

    说明：
    - 每项先做百分位排名（0~1）后加权
    - 对缺失项按“有效项重归一”处理，避免旧数据无新列时得分失真
    - 单个实验或缺失值按中性值 0.5 处理
    """
    scoring_items = [
        (COL_NAMES["daily_rankic_mean"], 0.30),
        (COL_NAMES["icir"], 0.30),
        (COL_NAMES["oos_top30_lift_mean"], 0.40),
    ]

    n = len(df)
    if n == 0:
        return pd.Series(dtype=float)
    if n == 1:
        return pd.Series(50.0, index=df.index)

    weighted_pct = pd.Series(0.0, index=df.index)
    effective_weight = pd.Series(0.0, index=df.index)

    for col, weight in scoring_items:
        if col not in df.columns:
            continue

        s = pd.to_numeric(df[col], errors="coerce")
        pct = s.rank(ascending=True, method="average", na_option="keep") / n
        pct = pct.fillna(0.5)

        valid_mask = s.notna().astype(float)
        weighted_pct += weight * pct
        effective_weight += weight * valid_mask

    # 对有效指标不足的行按中性分处理；其余按有效项权重重归一
    score = pd.Series(50.0, index=df.index)
    valid_rows = effective_weight > 0
    score.loc[valid_rows] = (weighted_pct.loc[valid_rows] / effective_weight.loc[valid_rows]) * 100
    return score.round(1)


def _cn_param_cols(keys: list[str], df: pd.DataFrame) -> list[str]:
    """将内部参数键转换为当前表中存在的中文列名。"""
    cols = []
    for key in keys:
        col = COL_NAMES.get(key, key)
        if col in df.columns and col not in cols:
            cols.append(col)
    return cols


def _comparison_model_param_cols(df: pd.DataFrame) -> list[str]:
    """对比报表中的模型参数默认忽略 seed 维度，避免重复试验被误判为不同超参。"""
    keys = [key for key in MODEL_PARAM_KEYS if key not in SEED_STABILITY_EXCLUDED_MODEL_KEYS]
    return _cn_param_cols(keys, df)


def _numeric_series(df: pd.DataFrame, col: str) -> pd.Series:
    if col not in df.columns:
        return pd.Series(np.nan, index=df.index, dtype=float)
    return pd.to_numeric(df[col], errors="coerce")


def _weighted_percentile_score(
    df: pd.DataFrame,
    scoring_items: list[tuple[str, float, str]],
) -> pd.Series:
    """对指定列做百分位加权评分，缺失指标按有效权重重归一。"""
    if df.empty:
        return pd.Series(dtype=float)
    if len(df) == 1:
        return pd.Series(50.0, index=df.index)

    weighted_pct = pd.Series(0.0, index=df.index)
    effective_weight = pd.Series(0.0, index=df.index)
    for col, weight, direction in scoring_items:
        if col not in df.columns:
            continue
        s = pd.to_numeric(df[col], errors="coerce")
        if direction == "abs_low":
            rank_input = s.abs()
            ascending = False
        elif direction == "low":
            rank_input = s
            ascending = False
        else:
            rank_input = s
            ascending = True
        pct = rank_input.rank(ascending=ascending, method="average", na_option="keep") / len(df)
        pct = pct.fillna(0.5)
        valid_mask = s.notna().astype(float)
        weighted_pct += weight * pct
        effective_weight += weight * valid_mask

    score = pd.Series(50.0, index=df.index)
    valid_rows = effective_weight > 0
    score.loc[valid_rows] = (weighted_pct.loc[valid_rows] / effective_weight.loc[valid_rows]) * 100
    return score.round(1)


def _signature_for_frame(df: pd.DataFrame, cols: list[str]) -> pd.Series:
    """为一组参数列生成可合并的稳定签名。"""
    if not cols:
        return pd.Series("", index=df.index)

    def normalize(value) -> str:
        if _is_missing_param_value(value):
            return "<NA>"
        return str(value).strip()

    return df[cols].apply(lambda row: "||".join(normalize(v) for v in row), axis=1)


def _unique_count(group: pd.DataFrame, cols: list[str]) -> int:
    if not cols:
        return len(group)
    return len(group[cols].drop_duplicates())


def _dedupe_columns(cols: list[str]) -> list[str]:
    result = []
    for col in cols:
        if col not in result:
            result.append(col)
    return result


def _build_model_alpha_score_table_for_cols(
    comp_df: pd.DataFrame,
    model_cols: list[str],
) -> pd.DataFrame:
    """按指定模型参数列聚合，构建只评价选股 alpha 的评分表。"""
    if comp_df.empty:
        return pd.DataFrame()

    if not model_cols:
        return pd.DataFrame()

    rows = []
    for _, group in comp_df.groupby(model_cols, dropna=False, sort=False):
        row = {col: group.iloc[0][col] for col in model_cols}
        period_col = COL_NAMES["batch_period_label"]
        row.update(
            {
                "样本数": len(group),
                "时间段数": (
                    int(group[period_col].dropna().nunique())
                    if period_col in group.columns
                    else None
                ),
                "运行ID列表": (
                    " | ".join(group[COL_NAMES["wf_run_id"]].astype(str).tolist())
                    if COL_NAMES["wf_run_id"] in group.columns
                    else None
                ),
                "选股综合得分均值": round(_numeric_series(group, "选股综合得分").mean(), 4),
                "选股综合得分最差": round(_numeric_series(group, "选股综合得分").min(), 4),
                "RankIC均值": round(
                    _numeric_series(group, COL_NAMES["daily_rankic_mean"]).mean(), 6
                ),
                "ICIR均值": round(_numeric_series(group, COL_NAMES["icir"]).mean(), 4),
                "Top30超额均值": round(
                    _numeric_series(group, COL_NAMES["oos_top30_lift_mean"]).mean(), 6
                ),
                "Top30胜率": round(
                    _numeric_series(group, COL_NAMES["oos_top30_win_rate"]).mean(), 4
                ),
                "Top30最差中位收益": round(
                    _numeric_series(group, COL_NAMES["oos_top30_worst_median"]).min(), 6
                ),
                "分层单调性均值": round(
                    _numeric_series(group, COL_NAMES["selection_monotonicity"]).mean(), 4
                ),
                "验证_OOS_IR差距": round(
                    _numeric_series(group, COL_NAMES["train_val_ir_gap"]).mean(), 4
                ),
                "全周期CAGR均值": round(_numeric_series(group, COL_NAMES["chain_cagr"]).mean(), 6),
                "全周期最大回撤最差": round(
                    _numeric_series(group, COL_NAMES["chain_max_drawdown"]).min(), 6
                ),
            }
        )
        rows.append(row)

    result = pd.DataFrame(rows)
    if result.empty:
        return result
    result["模型参数签名"] = _signature_for_frame(result, model_cols)
    result.insert(0, "模型Alpha分", _weighted_percentile_score(result, MODEL_ALPHA_SCORE_CONFIG))
    result.insert(
        1, "模型Alpha排名", result["模型Alpha分"].rank(ascending=False, method="min").astype(int)
    )
    result.insert(2, "模型参数组ID", [f"M{i:04d}" for i in range(1, len(result) + 1)])
    front_cols = [
        "模型Alpha分",
        "模型Alpha排名",
        "最新运行时间",
        "模型参数组ID",
        "样本数",
        "时间段数",
        "选股综合得分均值",
        "选股综合得分最差",
        "RankIC均值",
        "ICIR均值",
        "Top30超额均值",
        "Top30胜率",
        "Top30最差中位收益",
        "分层单调性均值",
        "验证_OOS_IR差距",
        "全周期CAGR均值",
        "全周期最大回撤最差",
        "运行ID列表",
        "模型参数签名",
    ]
    ordered = [col for col in front_cols if col in result.columns]
    ordered += [col for col in model_cols if col not in ordered]
    result = sort_by_latest_run_time(result, "运行ID列表", ["模型Alpha分"])
    ordered = [col for col in front_cols if col in result.columns]
    ordered += [col for col in model_cols if col not in ordered]
    return result[ordered].reset_index(drop=True)


def build_model_alpha_score_table(comp_df: pd.DataFrame) -> pd.DataFrame:
    """按非 seed 模型参数聚合，构建只评价选股 alpha 的评分表。"""
    model_cols = _comparison_model_param_cols(comp_df)
    return _build_model_alpha_score_table_for_cols(comp_df, model_cols)


def build_model_seed_stability_table(
    comp_df: pd.DataFrame,
    model_alpha_df: pd.DataFrame,
) -> pd.DataFrame:
    """按排除 seed 后的模型参数聚合，观察同一超参跨 seed 的稳定性。"""
    if comp_df.empty or model_alpha_df.empty:
        return pd.DataFrame()

    model_cols = _cn_param_cols(MODEL_PARAM_KEYS, comp_df)
    seed_cols = _cn_param_cols(SEED_STABILITY_EXCLUDED_MODEL_KEYS, comp_df)
    stable_model_cols = [col for col in model_cols if col not in seed_cols]
    if not stable_model_cols or not seed_cols:
        return pd.DataFrame()

    seed_level_alpha_df = _build_model_alpha_score_table_for_cols(comp_df, model_cols)
    if seed_level_alpha_df.empty:
        return pd.DataFrame()

    working = comp_df.copy()
    working["模型参数签名"] = _signature_for_frame(working, model_cols)
    working["Seed稳定性签名"] = _signature_for_frame(working, stable_model_cols)
    lookup_cols = [
        "模型参数签名",
        "模型参数组ID",
        "模型Alpha分",
        "运行ID列表",
    ]
    lookup = seed_level_alpha_df[
        [col for col in lookup_cols if col in seed_level_alpha_df.columns]
    ].drop_duplicates("模型参数签名")
    working = working.merge(lookup, on="模型参数签名", how="left", suffixes=("", "_模型Alpha"))

    rows = []
    for _, group in working.groupby(stable_model_cols, dropna=False, sort=False):
        alpha_by_model = group.drop_duplicates("模型参数签名")
        alpha_series = _numeric_series(alpha_by_model, "模型Alpha分")
        row = {col: group.iloc[0][col] for col in stable_model_cols}
        seed_values = []
        for col in seed_cols:
            values = group[col].dropna().astype(str).drop_duplicates().tolist()
            if values:
                seed_values.append(f"{col}=" + ",".join(values))
        run_lists = []
        run_list_col = (
            "运行ID列表_模型Alpha" if "运行ID列表_模型Alpha" in group.columns else "运行ID列表"
        )
        if run_list_col in group.columns:
            run_lists = group[run_list_col].dropna().astype(str).drop_duplicates().tolist()
        row.update(
            {
                "Seed稳定性样本数": int(alpha_by_model["模型参数签名"].nunique()),
                "Seed列表": " | ".join(seed_values),
                "模型Alpha分均值": (
                    round(alpha_series.mean(), 1) if alpha_series.notna().any() else None
                ),
                "模型Alpha分中位数": (
                    round(alpha_series.median(), 1) if alpha_series.notna().any() else None
                ),
                "模型Alpha分标准差": (
                    round(alpha_series.std(), 2) if alpha_series.notna().sum() > 1 else None
                ),
                "模型Alpha分最差": (
                    round(alpha_series.min(), 1) if alpha_series.notna().any() else None
                ),
                "模型Alpha分最好": (
                    round(alpha_series.max(), 1) if alpha_series.notna().any() else None
                ),
                "模型参数组ID列表": (
                    " | ".join(
                        alpha_by_model["模型参数组ID"]
                        .dropna()
                        .astype(str)
                        .drop_duplicates()
                        .tolist()
                    )
                    if "模型参数组ID" in alpha_by_model.columns
                    else None
                ),
                "运行ID列表": " | ".join(run_lists),
                "Seed稳定性签名": group["Seed稳定性签名"].iloc[0],
            }
        )
        rows.append(row)

    result = pd.DataFrame(rows)
    if result.empty:
        return result
    result["Seed稳健分"] = (
        _numeric_series(result, "模型Alpha分中位数") * 0.50
        + _numeric_series(result, "模型Alpha分最差") * 0.35
        + (100 - _numeric_series(result, "模型Alpha分标准差").fillna(0).clip(lower=0, upper=100))
        * 0.15
    ).round(1)
    result.insert(
        0, "Seed稳健排名", result["Seed稳健分"].rank(ascending=False, method="min").astype(int)
    )
    front_cols = [
        "Seed稳健分",
        "Seed稳健排名",
        "最新运行时间",
        "Seed稳定性样本数",
        "Seed列表",
        "模型Alpha分中位数",
        "模型Alpha分均值",
        "模型Alpha分标准差",
        "模型Alpha分最差",
        "模型Alpha分最好",
        "模型参数组ID列表",
        "运行ID列表",
        "Seed稳定性签名",
    ]
    result = sort_by_latest_run_time(result, "运行ID列表", ["Seed稳健分"])
    ordered = [col for col in front_cols if col in result.columns]
    ordered += [col for col in stable_model_cols if col not in ordered]
    return result[ordered].reset_index(drop=True)


def build_trade_param_score_table(comp_df: pd.DataFrame) -> pd.DataFrame:
    """在相同模型参数+相同时间段内做配对百分位，聚合交易参数评分。"""
    if comp_df.empty:
        return pd.DataFrame()

    model_cols = _comparison_model_param_cols(comp_df)
    context_cols = _cn_param_cols(PAIR_CONTEXT_KEYS, comp_df)
    trade_cols = _cn_param_cols(TRADE_PARAM_KEYS, comp_df)
    pair_cols = _dedupe_columns(
        [col for col in model_cols + context_cols if col in comp_df.columns]
    )
    if not pair_cols or not trade_cols:
        return pd.DataFrame()

    metric_map = [
        (COL_NAMES["chain_cagr"], "CAGR配对百分位", "high"),
        (COL_NAMES["chain_total_return"], "总收益配对百分位", "high"),
        (COL_NAMES["chain_sharpe"], "夏普配对百分位", "high"),
        (COL_NAMES["bt_calmar_mean"], "Calmar配对百分位", "high"),
        (COL_NAMES["bt_win_rate"], "胜率配对百分位", "high"),
        (COL_NAMES["chain_max_drawdown"], "最大回撤配对百分位", "high"),
    ]

    paired_frames = []
    for pair_index, (_, group) in enumerate(
        comp_df.groupby(pair_cols, dropna=False, sort=False), 1
    ):
        trade_candidate_count = _unique_count(group, trade_cols)
        if trade_candidate_count < 2:
            continue
        pair_df = group.copy()
        pair_df["__pair_context_id"] = f"P{pair_index:06d}"
        pair_df["__pair_candidate_count"] = trade_candidate_count
        for metric_col, pct_col, direction in metric_map:
            s = _numeric_series(pair_df, metric_col)
            ascending = direction == "high"
            pct = s.rank(ascending=ascending, method="average", na_option="keep") / len(pair_df)
            pair_df[pct_col] = pct.fillna(0.5) * 100
        pair_df["单次配对交易收益分"] = _weighted_percentile_score(
            pair_df,
            [
                ("CAGR配对百分位", 0.40, "high"),
                ("总收益配对百分位", 0.25, "high"),
                ("Calmar配对百分位", 0.15, "high"),
                ("夏普配对百分位", 0.10, "high"),
                ("胜率配对百分位", 0.10, "high"),
            ],
        )
        pair_df["单次配对交易稳健分"] = _weighted_percentile_score(
            pair_df,
            [
                ("最大回撤配对百分位", 0.35, "high"),
                ("Calmar配对百分位", 0.25, "high"),
                ("夏普配对百分位", 0.20, "high"),
                ("胜率配对百分位", 0.10, "high"),
            ],
        )
        paired_frames.append(pair_df)

    if not paired_frames:
        return pd.DataFrame(
            columns=["交易收益分", "交易稳健分", "有效配对环境数", "配对样本数"] + trade_cols
        )

    paired_df = pd.concat(paired_frames, ignore_index=False)
    paired_df["交易参数签名"] = _signature_for_frame(paired_df, trade_cols)
    rows = []
    for _, group in paired_df.groupby(trade_cols, dropna=False, sort=False):
        cagr_pair_min = group.groupby("__pair_context_id")["CAGR配对百分位"].min()
        row = {col: group.iloc[0][col] for col in trade_cols}
        row.update(
            {
                "有效配对环境数": int(group["__pair_context_id"].nunique()),
                "配对样本数": len(group),
                "平均每组候选数": round(group["__pair_candidate_count"].mean(), 2),
                "胜出率": round((group["单次配对交易收益分"] >= 50).mean(), 4),
                "交易收益分": round(group["单次配对交易收益分"].mean(), 1),
                "交易收益分标准差": (
                    round(group["单次配对交易收益分"].std(), 4) if len(group) > 1 else None
                ),
                "交易收益分最差": round(group["单次配对交易收益分"].min(), 1),
                "交易稳健分": round(group["单次配对交易稳健分"].mean(), 1),
                "CAGR配对百分位均值": round(group["CAGR配对百分位"].mean(), 2),
                "总收益配对百分位均值": round(group["总收益配对百分位"].mean(), 2),
                "夏普配对百分位均值": round(group["夏普配对百分位"].mean(), 2),
                "Calmar配对百分位均值": round(group["Calmar配对百分位"].mean(), 2),
                "胜率配对百分位均值": round(group["胜率配对百分位"].mean(), 2),
                "最大回撤配对百分位均值": round(group["最大回撤配对百分位"].mean(), 2),
                "CAGR最差配对百分位": (
                    round(cagr_pair_min.mean(), 2) if len(cagr_pair_min) else None
                ),
                "CAGR原始均值": round(_numeric_series(group, COL_NAMES["chain_cagr"]).mean(), 6),
                "CAGR原始最差": round(_numeric_series(group, COL_NAMES["chain_cagr"]).min(), 6),
                "最大回撤原始均值": round(
                    _numeric_series(group, COL_NAMES["chain_max_drawdown"]).mean(), 6
                ),
                "最大回撤原始最差": round(
                    _numeric_series(group, COL_NAMES["chain_max_drawdown"]).min(), 6
                ),
                "运行ID列表": (
                    " | ".join(group[COL_NAMES["wf_run_id"]].astype(str).tolist())
                    if COL_NAMES["wf_run_id"] in group.columns
                    else None
                ),
                "交易参数签名": group["交易参数签名"].iloc[0],
            }
        )
        rows.append(row)

    result = pd.DataFrame(rows)
    if result.empty:
        return result
    result.insert(
        1, "交易收益排名", result["交易收益分"].rank(ascending=False, method="min").astype(int)
    )
    result.insert(2, "交易参数组ID", [f"T{i:04d}" for i in range(1, len(result) + 1)])
    front_cols = [
        "交易收益分",
        "交易收益排名",
        "最新运行时间",
        "交易参数组ID",
        "交易稳健分",
        "有效配对环境数",
        "配对样本数",
        "平均每组候选数",
        "胜出率",
        "交易收益分标准差",
        "交易收益分最差",
        "CAGR配对百分位均值",
        "总收益配对百分位均值",
        "夏普配对百分位均值",
        "Calmar配对百分位均值",
        "胜率配对百分位均值",
        "最大回撤配对百分位均值",
        "CAGR最差配对百分位",
        "CAGR原始均值",
        "CAGR原始最差",
        "最大回撤原始均值",
        "最大回撤原始最差",
        "运行ID列表",
        "交易参数签名",
    ]
    ordered = [col for col in front_cols if col in result.columns]
    ordered += [col for col in trade_cols if col not in ordered]
    result = sort_by_latest_run_time(result, "运行ID列表", ["交易收益分"])
    ordered = [col for col in front_cols if col in result.columns]
    ordered += [col for col in trade_cols if col not in ordered]
    return result[ordered].reset_index(drop=True)


def build_live_candidate_score_table(
    comp_df: pd.DataFrame,
    model_score_df: pd.DataFrame,
    trade_score_df: pd.DataFrame,
) -> pd.DataFrame:
    """构建最终候选评分表，硬门槛未通过时实盘候选分直接置 0。"""
    if comp_df.empty or model_score_df.empty or trade_score_df.empty:
        return pd.DataFrame()

    model_cols = _comparison_model_param_cols(comp_df)
    trade_cols = _cn_param_cols(TRADE_PARAM_KEYS, comp_df)
    candidate_cols = _dedupe_columns(
        [col for col in model_cols + trade_cols if col in comp_df.columns]
    )
    if not candidate_cols:
        return pd.DataFrame()

    working = comp_df.copy()
    working["模型参数签名"] = _signature_for_frame(working, model_cols)
    working["交易参数签名"] = _signature_for_frame(working, trade_cols)

    model_lookup = model_score_df[["模型参数签名", "模型参数组ID", "模型Alpha分"]].drop_duplicates(
        "模型参数签名"
    )
    trade_lookup_cols = [
        "交易参数签名",
        "交易参数组ID",
        "交易收益分",
        "交易稳健分",
        "有效配对环境数",
    ]
    trade_lookup = trade_score_df[
        [c for c in trade_lookup_cols if c in trade_score_df.columns]
    ].drop_duplicates("交易参数签名")
    working = working.merge(model_lookup, on="模型参数签名", how="left")
    working = working.merge(trade_lookup, on="交易参数签名", how="left")

    rows = []
    group_cols = ["模型参数签名", "交易参数签名"]
    for _, group in working.groupby(group_cols, dropna=False, sort=False):
        row = {col: group.iloc[0][col] for col in candidate_cols}
        period_col = COL_NAMES["batch_period_label"]
        cagr_series = _numeric_series(group, COL_NAMES["chain_cagr"])
        drawdown_series = _numeric_series(group, COL_NAMES["chain_max_drawdown"])
        row.update(
            {
                "模型参数组ID": (
                    group["模型参数组ID"].iloc[0] if "模型参数组ID" in group.columns else None
                ),
                "交易参数组ID": (
                    group["交易参数组ID"].iloc[0] if "交易参数组ID" in group.columns else None
                ),
                "模型参数签名": group["模型参数签名"].iloc[0],
                "交易参数签名": group["交易参数签名"].iloc[0],
                "模型Alpha分": round(_numeric_series(group, "模型Alpha分").mean(), 1),
                "交易收益分": round(_numeric_series(group, "交易收益分").mean(), 1),
                "交易稳健分": round(_numeric_series(group, "交易稳健分").mean(), 1),
                "有效配对环境数": (
                    int(_numeric_series(group, "有效配对环境数").max())
                    if "有效配对环境数" in group.columns
                    and _numeric_series(group, "有效配对环境数").notna().any()
                    else 0
                ),
                "时间段数": (
                    int(group[period_col].dropna().nunique())
                    if period_col in group.columns
                    else None
                ),
                "样本数": len(group),
                "全周期CAGR均值": round(cagr_series.mean(), 6),
                "跨时间段CAGR最差": round(cagr_series.min(), 6),
                "全周期最大回撤最差": round(drawdown_series.min(), 6),
                "运行ID列表": (
                    " | ".join(group[COL_NAMES["wf_run_id"]].astype(str).tolist())
                    if COL_NAMES["wf_run_id"] in group.columns
                    else None
                ),
            }
        )
        rows.append(row)

    result = pd.DataFrame(rows)
    if result.empty:
        return result

    defense_input = result[["跨时间段CAGR最差", "全周期最大回撤最差"]].copy()
    result["最差场景防守分"] = _weighted_percentile_score(
        defense_input,
        [("跨时间段CAGR最差", 0.50, "high"), ("全周期最大回撤最差", 0.50, "high")],
    )
    result["实盘候选原始分"] = (
        _numeric_series(result, "模型Alpha分") * 0.45
        + _numeric_series(result, "交易收益分") * 0.30
        + _numeric_series(result, "交易稳健分") * 0.15
        + _numeric_series(result, "最差场景防守分") * 0.10
    ).round(1)

    model_pass = _numeric_series(result, "模型Alpha分") >= CANDIDATE_MIN_MODEL_ALPHA
    pair_pass = _numeric_series(result, "有效配对环境数") >= CANDIDATE_MIN_EFFECTIVE_PAIR_CONTEXTS
    drawdown_pass = (
        _numeric_series(result, "全周期最大回撤最差") >= CANDIDATE_MIN_CHAIN_MAX_DRAWDOWN
    )
    cagr_pass = _numeric_series(result, "跨时间段CAGR最差") >= CANDIDATE_MIN_CHAIN_CAGR_WORST
    result["模型Alpha门槛通过"] = model_pass
    result["有效配对门槛通过"] = pair_pass
    result["最大回撤门槛通过"] = drawdown_pass
    result["最差CAGR门槛通过"] = cagr_pass
    result["候选门槛通过"] = model_pass & pair_pass & drawdown_pass & cagr_pass

    failure_reasons = []
    for _, row in result.iterrows():
        reasons = []
        if not bool(row["模型Alpha门槛通过"]):
            reasons.append(f"模型Alpha分<{CANDIDATE_MIN_MODEL_ALPHA:g}")
        if not bool(row["有效配对门槛通过"]):
            reasons.append(f"有效配对环境数<{CANDIDATE_MIN_EFFECTIVE_PAIR_CONTEXTS}")
        if not bool(row["最大回撤门槛通过"]):
            reasons.append(f"全周期最大回撤<{CANDIDATE_MIN_CHAIN_MAX_DRAWDOWN:.0%}")
        if not bool(row["最差CAGR门槛通过"]):
            reasons.append(f"跨时间段CAGR最差<{CANDIDATE_MIN_CHAIN_CAGR_WORST:.0%}")
        failure_reasons.append("；".join(reasons) if reasons else "")
    result["候选门槛失败原因"] = failure_reasons
    result["实盘候选分"] = result["实盘候选原始分"].where(result["候选门槛通过"], 0.0)
    result.insert(
        1, "候选排名", result["实盘候选分"].rank(ascending=False, method="min").astype(int)
    )

    front_cols = [
        "实盘候选分",
        "候选排名",
        "最新运行时间",
        "实盘候选原始分",
        "候选门槛通过",
        "候选门槛失败原因",
        "模型Alpha分",
        "交易收益分",
        "交易稳健分",
        "最差场景防守分",
        "有效配对环境数",
        "模型Alpha门槛通过",
        "有效配对门槛通过",
        "最大回撤门槛通过",
        "最差CAGR门槛通过",
        "模型参数组ID",
        "交易参数组ID",
        "时间段数",
        "样本数",
        "全周期CAGR均值",
        "跨时间段CAGR最差",
        "全周期最大回撤最差",
        "运行ID列表",
        "模型参数签名",
        "交易参数签名",
    ]
    ordered = [col for col in front_cols if col in result.columns]
    ordered += [col for col in candidate_cols if col not in ordered]
    result = sort_by_latest_run_time(result, "运行ID列表", ["实盘候选分"])
    ordered = [col for col in front_cols if col in result.columns]
    ordered += [col for col in candidate_cols if col not in ordered]
    return result[ordered].reset_index(drop=True)


def build_metric_descriptions() -> pd.DataFrame:
    """构建指标说明表（第二个 sheet）"""
    rows = [
        # ── 综合评分 ──────────────────────────────────────────────────────────
        (
            "综合评分",
            "综合得分",
            "跨实验百分位排名加权综合得分（0~100，越高越好）。"
            "对12个关键指标分别计算当前实验集内的百分位排名（0~1），再按权重求和×100。"
            "权重配置（回测60%）：全周期CAGR 20%、回测胜率 15%、回测夏普 15%、全周期链式最大回撤 10%；"
            "（统计32%）：跨切分IR 10%、Top30胜率 5%、Top30最差中位收益 5%、"
            "RankIC_IR趋势 5%、Top30中位收益 3%、Top30超额 2%、偏斜度 2%（绝对值低好）；"
            "（训练质量8%）：验证_OOS_IR差距 8%（低好）。"
            "NaN指标以中性分（0.5百分位）计入；仅1组实验时固定得50分。",
            "越高越好",
        ),
        (
            "综合评分",
            "选股综合得分",
            "纯选股能力评分（0~100，越高越好）。"
            "基于3项最核心选股指标做百分位加权：RankIC均值30%、ICIR30%、Top30超额均值40%。"
            "RankIC均值衡量排序方向，ICIR衡量排序稳定性，Top30超额均值衡量买入头部能否跑赢全市场。"
            "对缺失指标按有效项重归一；若该行全部缺失则记为50分。",
            "越高越好",
        ),
        (
            "三视角评分",
            "模型Alpha分",
            "按模型参数聚合后的纯选股 Alpha 评分（0~100，越高越好）。"
            "权重：选股综合得分均值30%、ICIR均值20%、Top30超额均值15%、Top30最差中位收益15%、"
            "分层单调性10%、验证_OOS_IR差距10%（低好）。用于回答“哪个模型/超参更好”。",
            "越高越好",
        ),
        (
            "三视角评分",
            "交易收益分",
            "按交易参数聚合后的收益评分（0~100，越高越好）。"
            "先在相同模型参数+相同时间段的配对环境内比较不同交易参数，计算收益类百分位，"
            "再跨模型、跨时间段取平均。单候选配对环境不参与评分。",
            "越高越好",
        ),
        (
            "三视角评分",
            "交易稳健分",
            "按交易参数聚合后的风控稳健评分（0~100，越高越好）。"
            "更重视最大回撤、Calmar、夏普、胜率和最差 CAGR 配对表现，用于辅助交易参数筛选。",
            "越高越好",
        ),
        (
            "三视角评分",
            "实盘候选分",
            "模型 Alpha、交易收益、交易稳健和最差场景防守的组合候选分。"
            "硬门槛未通过时直接置 0：模型Alpha分>=60、有效配对环境数>=2、"
            "全周期最大回撤最差>=-35%、跨时间段CAGR最差>=-5%。",
            "越高越好；0分优先看失败原因",
        ),
        (
            "三视角评分",
            "候选门槛失败原因",
            "实盘候选分被置 0 的具体原因，可能包含模型Alpha不足、有效配对环境不足、"
            "最大回撤过深或最差时间段 CAGR 过低。",
            "为空表示通过",
        ),
        (
            "Seed稳定性",
            "Seed稳健分",
            "忽略多种子字段后，对同一套模型超参跨 seed 的稳健评分。"
            "权重：模型Alpha分均值50%、模型Alpha分最差35%、模型Alpha分标准差惩罚15%。",
            "越高越好",
        ),
        (
            "Seed稳定性",
            "模型Alpha分均值",
            "同一套非 seed 模型超参下，不同 seed 对应模型Alpha分的平均值。",
            "越高越好",
        ),
        (
            "Seed稳定性",
            "模型Alpha分标准差",
            "同一套非 seed 模型超参下，不同 seed 对应模型Alpha分的标准差；越大说明 seed 敏感性越强。",
            "越低越稳",
        ),
        (
            "Seed稳定性",
            "模型Alpha分最差",
            "同一套非 seed 模型超参下，不同 seed 对应模型Alpha分的最差值，用于识别是否只靠某个 seed 碰巧跑好。",
            "越高越好",
        ),
        (
            "Seed稳定性",
            "Seed稳定性样本数",
            "同一套非 seed 模型超参下参与聚合的完整模型参数组数量，通常等于不同 seed 配置数量。",
            "越多越可靠",
        ),
        # ── OOS 性能指标 ──────────────────────────────────────────────────────
        (
            "OOS性能",
            "运行ID",
            "walk-forward运行的唯一标识符，格式为wf_YYYYMMDD_HHMMSS_xxxxxxxx",
            "标识符，无优劣",
        ),
        ("OOS性能", "切分数", "本次实验成功完成的OOS切分数量，越多统计结论越可靠", "越多越好"),
        (
            "OOS性能",
            "模型版本范围",
            "本次walk-forward生成的模型编号范围（格式：最小编号~最大编号），可在ModelRegistry中定位具体模型文件",
            "参考",
        ),
        (
            "OOS性能",
            "OOS_RankIC_IR均值",
            "各切分OOS期逐日RankIC信息比率（均值/标准差）的跨切分均值，衡量预测对股票排序的整体有效性",
            "越高越好",
        ),
        (
            "OOS性能",
            "OOS_RankIC_IR标准差",
            "各切分OOS RankIC IR的标准差，衡量策略在不同时间段的稳定性",
            "越低越稳定",
        ),
        (
            "OOS性能",
            "跨切分IR",
            "OOS_RankIC_IR均值 / 标准差，类似夏普比率，同时衡量收益水平与跨时间段稳定性，是排序各实验的首要指标",
            "越高越好（首要排序指标）",
        ),
        (
            "OOS性能",
            "RankIC均值",
            "各切分 OOS 逐日RankIC均值的跨切分均值；直接衡量排序相关性的绝对水平",
            "越高越好",
        ),
        ("OOS性能", "ICIR", "RankIC均值 / RankIC标准差，衡量单位波动下的排序信息效率", "越高越好"),
        (
            "OOS性能",
            "分层单调性(近似)",
            "基于 Top30/Top100/Top300 中位收益构造的单调性评分（0~1）；若收益随覆盖范围扩大递减，则得分更高",
            "越高越好",
        ),
        (
            "OOS性能",
            "RankIC_IR趋势(近-早)",
            "最近3个切分的IR均值 - 最早3个切分的IR均值；正值说明模型随时间改善，负值说明alpha在衰减",
            "接近0或正值为好，持续负值需警惕",
        ),
        (
            "OOS性能",
            "Top30中位收益均值",
            "各切分中每日Top30持仓20日收益中位数的跨切分均值；用中位数代替均值，减少极端行情日（如大涨停日）的干扰",
            "越高越好",
        ),
        (
            "OOS性能",
            "Top30胜率",
            "各切分中Top30中位收益>0的占比；衡量策略在不同历史时段的正收益稳健性，>70%可认为优秀",
            "越高越好（>0.7为优秀）",
        ),
        (
            "OOS性能",
            "Top30最差中位收益",
            "所有切分中Top30中位收益的最小值，代表策略的最差历史表现，用于压力测试",
            "越高越好（大幅负值需警惕）",
        ),
        (
            "OOS性能",
            "Top30偏斜度均值",
            "各切分中(Top30均值-中位数)/标准差的均值；偏斜度高说明均值被少数极端行情日拉偏，均值的代表性下降",
            "越接近0越好（>0.6需警惕）",
        ),
        (
            "OOS性能",
            "Top30超额均值",
            "各切分中Top30相对全市场平均收益的超额均值，衡量纯选股能力（剔除市场整体涨跌的影响）",
            "越高越好",
        ),
        (
            "OOS性能",
            "Top100中位收益均值",
            "各切分每日Top100持仓20日收益中位数的跨切分均值",
            "越高越好",
        ),
        ("OOS性能", "Top100胜率", "各切分中Top100中位收益>0的占比", "越高越好"),
        (
            "OOS性能",
            "Top300中位收益均值",
            "各切分每日Top300持仓20日收益中位数的跨切分均值；样本量大，统计更稳定但个股alpha被稀释",
            "越高越好",
        ),
        ("OOS性能", "Top300胜率", "各切分中Top300中位收益>0的占比", "越高越好"),
        # ── OOS 回测指标 ──────────────────────────────────────────────────────
        (
            "OOS回测",
            "全周期CAGR",
            "基于 chain_nav 将所有 split 的 OOS 净值顺次串联后得到的全周期复利年化收益，更适合做最终策略筛选",
            "越高越好（优先看）",
        ),
        (
            "OOS回测",
            "全周期总收益",
            "基于 chain_nav 串联后的起止总收益率，反映整轮 walk-forward 结束时的真实累计收益",
            "越高越好",
        ),
        (
            "OOS回测",
            "全周期链式最大回撤",
            "基于 chain_nav 串联后的全周期最大回撤，避免只看单个 split 的最差值而忽略跨 split 累计损失",
            "越接近0越好（优先看）",
        ),
        (
            "OOS回测",
            "全周期链式夏普",
            "基于 chain_nav 串联后的全周期夏普比率，使用 3% 无风险利率与全周期日收益波动计算",
            "越高越好",
        ),
        (
            "OOS回测",
            "全周期链式交易日数",
            "chain_nav 串联后的交易日总数，用于判断 CAGR 计算口径和样本长度是否一致",
            "参考",
        ),
        (
            "OOS回测",
            "回测年化收益均值",
            "各切分 OOS 回测年化收益率的跨切分均值，适合做横向粗筛，但不等同于全周期 CAGR",
            "越高越好（粗筛用）",
        ),
        (
            "OOS回测",
            "回测夏普均值",
            "各切分OOS回测夏普比率（年化收益-3%无风险利率/年化波动率）的跨切分均值",
            "越高越好（>1.0为优秀）",
        ),
        (
            "OOS回测",
            "回测最大回撤(最差)",
            "所有切分 OOS 回测中最大回撤的最差值（绝对值最大的单段回撤），用于观察最脆弱 split，但不等同于全周期链式回撤",
            "越接近0越好（-30%以下需警惕）",
        ),
        (
            "OOS回测",
            "回测Calmar均值",
            "各切分年化收益/最大回撤的均值，衡量单位风险回报",
            "越高越好（>1.0为良好）",
        ),
        (
            "OOS回测",
            "回测胜率",
            "各切分 OOS 回测总收益>0 的占比，衡量策略在不同历史时段的盈利稳健性",
            "越高越好（>0.7为优秀）",
        ),
        (
            "OOS回测",
            "回测总收益均值",
            "各切分 OOS 回测期间总收益率的跨切分均值，适合观察 split 层面的平均水平，不等同于全周期累计收益",
            "越高越好",
        ),
        ("OOS回测", "回测波动率均值", "各切分OOS回测年化波动率的跨切分均值", "越低越稳定"),
        # ── 训练质量指标 ──────────────────────────────────────────────────────
        (
            "训练质量",
            "验证集RankIC_IR均值",
            "各切分内部验证集逐日RankIC IR的跨切分均值；验证集来自训练窗口末尾，反映模型在样本内末期的泛化能力",
            "越高越好",
        ),
        (
            "训练质量",
            "验证_OOS_IR差距",
            "验证集IR均值 - OOS IR均值；正值表示验证集优于OOS（轻度过拟合信号），负值表示OOS优于验证集（通常正常）",
            "接近0为好，负值可接受，大正值（>0.5）需警惕过拟合",
        ),
        (
            "训练质量",
            "最佳迭代均值",
            "各切分早停触发时的迭代次数均值；反映模型实际使用的树数量，可指导n_estimators上限的设置",
            "参考指标（<100可能欠拟合，接近n_estimators上限则建议增大）",
        ),
        (
            "训练质量",
            "最佳迭代最小值",
            "所有切分中最少的早停迭代次数；如果某个切分迭代极少，说明该时段数据可能有异常",
            "不宜过低（<100需关注）",
        ),
        (
            "训练质量",
            "最佳迭代最大值",
            "所有切分中最多的早停迭代次数",
            "不宜接近n_estimators（说明需增大树数量上限）",
        ),
        (
            "训练质量",
            "最佳迭代标准差",
            "各切分迭代次数的标准差；衡量模型在不同时间段需要的学习量是否一致，差异过大说明数据分布在各时段差异明显",
            "越低越稳定",
        ),
        # ── 训练参数（仅供参考） ───────────────────────────────────────────────
        ("训练参数", "WF起始日期", "walk-forward整体时间范围起始日期，训练集不早于此日期", "参考"),
        (
            "训练参数",
            "WF结束日期",
            "walk-forward整体时间范围结束日期，训练集截止于此，测试集可超出此范围",
            "参考",
        ),
        (
            "训练参数",
            "滚动频率",
            "每次train_end向前推进的步长（monthly=月度 / quarterly=季度 / semiannual=半年度）",
            "参考",
        ),
        (
            "训练参数",
            "训练窗口年数",
            "每次训练使用的历史数据年数；过短欠拟合，过长可能纳入失效的历史规律",
            "参考（通常3~7年）",
        ),
        (
            "训练参数",
            "测试窗口月数",
            "每次OOS评估的时间窗口（月数），建议与标签持仓周期相近",
            "参考",
        ),
        ("训练参数", "验证集比例", "训练数据中用于内部早停评估的比例", "参考"),
        (
            "训练参数",
            "标签列",
            "预测目标列名，如neu_y_ret_20（中性化20日收益）；neu_前缀表示已剔除市值/行业因子",
            "参考",
        ),
        (
            "训练参数",
            "任务类型",
            "regression=回归（预测收益率大小）/ classification=分类（预测涨跌方向）",
            "参考",
        ),
        (
            "训练参数",
            "标签变换",
            "raw=使用原始收益率标签 / cs_zscore=截面z-score标准化（消除截面异方差，让模型聚焦排序）",
            "参考",
        ),
        (
            "训练参数",
            "树数量",
            "XGBoost决策树总数上限；配合早停使用，实际用量为最佳迭代次数",
            "参考",
        ),
        (
            "训练参数",
            "最大深度",
            "每棵树的最大层数；越大模型越复杂越容易过拟合",
            "参考（建议6~10）",
        ),
        ("训练参数", "学习率", "梯度下降步长；越小越精细但需要更多树才能收敛", "参考"),
        ("训练参数", "样本采样比", "每棵树随机抽取的样本比例（行采样），降低过拟合风险", "参考"),
        ("训练参数", "特征采样比", "每棵树随机抽取的特征比例（列采样），降低过拟合风险", "参考"),
        (
            "训练参数",
            "最小叶节点权重",
            "叶节点所需的最少样本权重之和；越大越保守，防止模型学习噪声",
            "参考",
        ),
        ("训练参数", "gamma", "节点分裂所需的最小损失下降量；越大越保守，减少无效分裂", "参考"),
        ("训练参数", "L1正则", "L1正则化系数，使特征权重趋向稀疏（部分特征权重归零）", "参考"),
        ("训练参数", "L2正则", "L2正则化系数，使特征权重趋向平滑（防止某个特征权重过大）", "参考"),
        (
            "训练参数",
            "rank权重启用",
            "是否对每日Top/Bottom K样本赋予更高训练权重，使模型更关注极端收益样本",
            "参考",
        ),
        ("训练参数", "rank权重TopK", "每日增强权重覆盖的头部/尾部股票数量", "参考"),
        ("训练参数", "rank权重值", "增强样本相对普通样本的权重倍数", "参考"),
    ]
    return pd.DataFrame(rows, columns=["分类", "指标名", "说明", "优劣方向"])


def build_split_detail_table(all_df: pd.DataFrame) -> pd.DataFrame:
    """构建逐 split 明细表（每个 split 一行，展示该 split 的回测表现与训练参数）

    用于对比不同实验在各时间段的表现差异，帮助定位哪些时段拖后腿。
    """
    if all_df.empty or "wf_run_id" not in all_df.columns:
        return pd.DataFrame()

    # 选取需要的列
    detail_cols = [
        "wf_run_id",
        "split_index",
        "KEY_Top20_list",
        "KEY_Top30_list",
        "KEY_Top20_hit_rate",
        "KEY_Top20_avg_return_median",
        "KEY_Top30_hit_rate",
        "KEY_Top30_avg_return_median",
        "train_start",
        "train_end",
        "test_start",
        "test_end",
        # 逐 split 回测指标
        "bt_total_return",
        "bt_annual_return",
        "bt_max_drawdown",
        "bt_sharpe",
        "bt_calmar",
        "bt_volatility",
        "bt_trading_days",
        "bt_top_n",
        "bt_rebalance_freq",
        # 逐 split 模型质量
        "daily_rankic_mean",
        "daily_rankic_ir",
        "val_rankic_ir",
        "best_iteration",
        "train_samples",
        "test_samples",
        # 关键训练参数（用于区分实验）
        "algorithm",
        "label_column",
        "task",
        "label_transform",
        "max_depth",
        "learning_rate",
        "train_window_years",
        "test_window_months",
        "time_decay_half_life",
        "market_regime",
        "enable_fundamental",
    ]
    available_cols = [c for c in detail_cols if c in all_df.columns]
    result = all_df[available_cols].copy()

    # 按运行时间降序、split 升序排列
    result = result.sort_values(["wf_run_id", "split_index"], ascending=[False, True])

    # 计算累计净值（组内 cumprod）
    if "bt_total_return" in result.columns:
        result["chain_nav"] = result.groupby("wf_run_id")["bt_total_return"].transform(
            lambda x: (1 + x).cumprod()
        )

    # 列名翻译
    rename_map = {
        "wf_run_id": "运行ID",
        "split_index": "切分序号",
        "KEY_Top20_list": "重点Top20名单",
        "KEY_Top30_list": "重点Top30名单",
        "KEY_Top20_hit_rate": "重点Top20命中率",
        "KEY_Top20_avg_return_median": "重点Top20收益中位数",
        "KEY_Top30_hit_rate": "重点Top30命中率",
        "KEY_Top30_avg_return_median": "重点Top30收益中位数",
        "train_start": "训练开始",
        "train_end": "训练结束",
        "test_start": "测试开始",
        "test_end": "测试结束",
        "bt_total_return": "总收益",
        "bt_annual_return": "年化收益",
        "bt_max_drawdown": "最大回撤",
        "bt_sharpe": "夏普",
        "bt_calmar": "Calmar",
        "bt_volatility": "波动率",
        "bt_trading_days": "交易天数",
        "bt_top_n": "TopN",
        "bt_rebalance_freq": "调仓频率",
        "daily_rankic_mean": "OOS_RankIC均值",
        "daily_rankic_ir": "OOS_RankIC_IR",
        "val_rankic_ir": "验证集RankIC_IR",
        "best_iteration": "最佳迭代",
        "train_samples": "训练样本数",
        "test_samples": "测试样本数",
        "chain_nav": "累计净值",
        "algorithm": "算法",
        "label_column": "标签列",
        "task": "任务类型",
        "label_transform": "标签变换",
        "max_depth": "最大深度",
        "learning_rate": "学习率",
        "train_window_years": "训练窗口年数",
        "test_window_months": "测试窗口月数",
        "time_decay_half_life": "时间衰减半衰期",
        "market_regime": "市场择时",
        "enable_fundamental": "基本面因子",
    }
    result = result.rename(columns={k: v for k, v in rename_map.items() if k in result.columns})

    # 调整列序：运行ID → 切分序号 → 时间 → 回测指标 → 累计净值 → 模型质量 → 训练参数
    ordered = [
        "运行ID",
        "切分序号",
        "重点Top20命中率",
        "重点Top20收益中位数",
        "重点Top30命中率",
        "重点Top30收益中位数",
        "重点Top20名单",
        "重点Top30名单",
        "训练开始",
        "训练结束",
        "测试开始",
        "测试结束",
        "总收益",
        "年化收益",
        "最大回撤",
        "夏普",
        "Calmar",
        "波动率",
        "交易天数",
        "TopN",
        "累计净值",
        "OOS_RankIC均值",
        "OOS_RankIC_IR",
        "验证集RankIC_IR",
        "最佳迭代",
        "训练样本数",
        "测试样本数",
        "算法",
        "标签列",
        "任务类型",
        "标签变换",
        "最大深度",
        "学习率",
        "训练窗口年数",
        "测试窗口月数",
        "时间衰减半衰期",
        "市场择时",
        "基本面因子",
    ]
    final_cols = [c for c in ordered if c in result.columns]
    return result[final_cols].reset_index(drop=True)


def sort_by_run_time(df: pd.DataFrame, run_id_col: str = "运行ID") -> pd.DataFrame:
    """按 wf_run_id 中的运行时间戳降序排列（最近的排在前面）

    wf_run_id 格式: wf_YYYYMMDD_HHMMSS_xxxxxxxx
    提取 YYYYMMDD_HHMMSS 作为排序键。
    """
    return sort_by_latest_run_time(df, run_id_col, ["综合得分"])


def reorder_comparison_columns(df: pd.DataFrame) -> pd.DataFrame:
    """按汇总阅读习惯重排实验对比列顺序。"""
    preferred_order = [
        "运行ID",
        "最新运行时间",
        "综合得分",
        "选股综合得分",
        "RankIC均值",
        "ICIR",
        "Top30超额均值",
        "最大深度",
        "学习率",
        "rank权重TopK",
        "rank权重值",
    ]

    tail_order = [
    ]

    ordered_cols = [col for col in preferred_order if col in df.columns]

    # 其余列保持原顺序，避免非目标列发生意外位移。
    for col in df.columns:
        if col not in ordered_cols and col not in tail_order:
            ordered_cols.append(col)

    ordered_cols.extend([col for col in tail_order if col in df.columns])
    return df.reindex(columns=ordered_cols)


def _series_has_meaningful_variation(series: pd.Series) -> bool:
    """判断列是否存在实际区分度，纯常量或仅空值视为无区分度。"""
    if series is None:
        return False
    non_na = series.dropna()
    if non_na.empty:
        return False
    return non_na.astype(str).nunique() > 1


def compact_experiment_sheet_for_display(df: pd.DataFrame, source_label: str) -> pd.DataFrame:
    """压缩 batches 实验对比页，只保留核心指标和有区分度的关键参数。"""
    if df.empty or source_label != "batches":
        return df

    ordered: list[str] = []
    for col in BATCH_EXPERIMENT_CORE_COLS:
        if col in df.columns and col not in ordered:
            ordered.append(col)

    for col in BATCH_EXPERIMENT_PARAM_CANDIDATES:
        if (
            col in df.columns
            and col not in ordered
            and col not in BATCH_EXPERIMENT_EXCLUDED_PARAM_COLS
            and _series_has_meaningful_variation(df[col])
        ):
            ordered.append(col)

    return df[[col for col in ordered if col in df.columns]].copy()


def _str_display_width(s: str) -> int:
    """计算字符串的显示宽度（CJK 字符计为 2，ASCII 计为 1）"""
    width = 0
    for ch in str(s):
        eaw = unicodedata.east_asian_width(ch)
        width += 2 if eaw in ("W", "F") else 1
    return width


def _weight_to_green_fill(weight: float, max_weight: float) -> PatternFill:
    """将权重映射为浅绿→中绿的填充色（权重越高越深，黑字始终可读）

    颜色范围：
      最小权重 → RGB(220, 245, 220)  极浅绿
      最大权重 → RGB(130, 215, 130)  中等绿（与黑字对比度 ≈ 8:1，远超 WCAG AA 4.5:1）
    """
    t = min(weight / max_weight, 1.0) if max_weight > 0 else 0.0
    r = int(round(220 - t * 90))  # 220 → 130
    g = int(round(245 - t * 30))  # 245 → 215
    b = int(round(220 - t * 90))  # 220 → 130
    return PatternFill(fill_type="solid", fgColor=f"{r:02X}{g:02X}{b:02X}")


def _weight_to_palette_fill(weight: float, max_weight: float, palette: str) -> PatternFill:
    """将权重映射为指定色系的浅色填充，权重越高颜色越深。"""
    if palette == "blue":
        start, end = (225, 239, 255), (132, 181, 232)
    elif palette == "orange":
        start, end = (255, 235, 205), (242, 174, 88)
    elif palette == "red":
        start, end = (255, 224, 224), (242, 150, 150)
    else:
        start, end = (220, 245, 220), (130, 215, 130)
    t = min(weight / max_weight, 1.0) if max_weight > 0 else 0.0
    r = int(round(start[0] + (end[0] - start[0]) * t))
    g = int(round(start[1] + (end[1] - start[1]) * t))
    b = int(round(start[2] + (end[2] - start[2]) * t))
    return PatternFill(fill_type="solid", fgColor=f"{r:02X}{g:02X}{b:02X}")


def _score_column_fills() -> dict[str, dict[str, PatternFill]]:
    """构建各评分 sheet 的列填充配置。"""
    max_model = max(w for _, w, _ in MODEL_ALPHA_SCORE_CONFIG)
    max_trade = max(w for _, w, _ in TRADE_YIELD_SCORE_CONFIG)
    max_candidate = max(w for _, w, _ in CANDIDATE_SCORE_CONFIG)
    return {
        "模型Alpha评分": {
            "模型Alpha分": _weight_to_palette_fill(max_model, max_model, "blue"),
            **{
                col: _weight_to_palette_fill(weight, max_model, "blue")
                for col, weight, _ in MODEL_ALPHA_SCORE_CONFIG
            },
        },
        "模型Seed稳定性": {
            "Seed稳健分": _weight_to_palette_fill(max_model, max_model, "blue"),
            "模型Alpha分均值": _weight_to_palette_fill(0.30, max_model, "blue"),
            "模型Alpha分标准差": _weight_to_palette_fill(0.15, max_model, "blue"),
            "模型Alpha分最差": _weight_to_palette_fill(0.25, max_model, "blue"),
            "模型Alpha分最好": _weight_to_palette_fill(0.10, max_model, "blue"),
        },
        "交易参数收益评分": {
            "交易收益分": _weight_to_palette_fill(max_trade, max_trade, "orange"),
            "交易稳健分": _weight_to_palette_fill(max_trade, max_trade, "orange"),
            **{
                col: _weight_to_palette_fill(weight, max_trade, "orange")
                for col, weight, _ in TRADE_YIELD_SCORE_CONFIG
            },
        },
        "实盘候选评分": {
            "实盘候选分": _weight_to_palette_fill(max_candidate, max_candidate, "green"),
            "实盘候选原始分": _weight_to_palette_fill(max_candidate, max_candidate, "green"),
            **{
                col: _weight_to_palette_fill(weight, max_candidate, "green")
                for col, weight, _ in CANDIDATE_SCORE_CONFIG
            },
        },
    }


def format_excel_output(wb, desc_df: pd.DataFrame) -> None:
    """对 Excel 工作簿应用格式化

    - 全局字体: 微软雅黑 9 号
    - 冻结标题行（两个 sheet 均适用）
    - 实验对比标题行超链接跳转至指标说明对应行
    - 自动列宽（CJK 双倍宽）
    - 参与综合得分的列着浅绿背景，权重越高越深
    """
    # 构建 指标名 → 指标说明 sheet 行号的映射（第 1 行为标题，数据从第 2 行起）
    desc_row_map: dict[str, int] = {}
    for i, row in desc_df.iterrows():
        desc_row_map[str(row["指标名"])] = int(i) + 2  # +2: header占第1行，数据从第2行

    font_normal = Font(name="微软雅黑", size=9)
    font_link = Font(name="微软雅黑", size=9, color="0563C1", underline="single")

    # ── 构建 中文列名 → 填充 的映射（用于各评分 sheet）─────────────────
    _max_w = max(w for _, w, _ in SCORE_CONFIG) if SCORE_CONFIG else 1.0
    # 综合得分列本身也着色，用最深绿（权重等同最大权重）
    score_cn_fills: dict[str, PatternFill] = {
        "综合得分": _weight_to_green_fill(_max_w, _max_w),
    }
    for eng_key, weight, _ in SCORE_CONFIG:
        col_cn = COL_NAMES.get(eng_key)
        if col_cn:
            score_cn_fills[col_cn] = _weight_to_green_fill(weight, _max_w)

    sheet_score_fills = _score_column_fills()
    key_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    zero_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")
    score_header_fill = PatternFill(fill_type="solid", fgColor="274E13")
    score_header_font = Font(name="微软雅黑", size=9, color="FFFFFF", bold=True)
    key_cols_cn = {
        "重点说明",
        "重点Top20最新名单",
        "重点Top30最新名单",
        "重点Top20命中率均值",
        "重点Top20收益中位数均值",
        "重点Top30命中率均值",
        "重点Top30收益中位数均值",
    }

    sheet_col_letter_fill: dict[str, dict[str, PatternFill]] = {}
    for sheet_name in wb.sheetnames:
        ws_ref = wb[sheet_name]
        col_letter_fill: dict[str, PatternFill] = {}
        local_fills = (
            score_cn_fills if sheet_name == "实验对比" else sheet_score_fills.get(sheet_name, {})
        )
        for cell in next(ws_ref.iter_rows(min_row=1, max_row=1)):
            if cell.value and str(cell.value) in local_fills:
                col_letter_fill[cell.column_letter] = local_fills[str(cell.value)]
            elif sheet_name == "实验对比" and cell.value and str(cell.value) in key_cols_cn:
                col_letter_fill[cell.column_letter] = key_fill
        sheet_col_letter_fill[sheet_name] = col_letter_fill

    # ── 全局字体、冻结、列宽、绿色背景 ──────────────────────────────────
    all_sheets = [
        s
        for s in [
            "实盘候选评分",
            "模型Alpha评分",
            "模型Seed稳定性",
            "交易参数收益评分",
            "实验对比",
            "跨时间段稳定性",
            "指标说明",
            "逐Split明细",
        ]
        if s in wb.sheetnames
    ]
    for sheet_name in all_sheets:
        ws = wb[sheet_name]
        ws.freeze_panes = "A2"
        header_values = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        zero_score_col = None
        failure_cols = set()
        if sheet_name == "实盘候选评分":
            for idx, value in enumerate(header_values, 1):
                if value == "实盘候选分":
                    zero_score_col = idx
                if value in {
                    "候选门槛失败原因",
                    "候选门槛通过",
                    "模型Alpha门槛通过",
                    "有效配对门槛通过",
                    "最大回撤门槛通过",
                    "最差CAGR门槛通过",
                }:
                    failure_cols.add(idx)

        col_widths: dict[str, int] = {}
        for row in ws.iter_rows():
            is_zero_candidate = False
            if sheet_name == "实盘候选评分" and row[0].row > 1 and zero_score_col is not None:
                score_value = row[zero_score_col - 1].value
                try:
                    is_zero_candidate = float(score_value) == 0.0
                except (TypeError, ValueError):
                    is_zero_candidate = False
            for cell in row:
                cell.font = font_normal
                if cell.row == 1 and cell.column_letter in sheet_col_letter_fill.get(
                    sheet_name, {}
                ):
                    cell.fill = score_header_fill
                    cell.font = score_header_font
                elif cell.column_letter in sheet_col_letter_fill.get(sheet_name, {}):
                    cell.fill = sheet_col_letter_fill[sheet_name][cell.column_letter]
                if is_zero_candidate and (
                    cell.column == zero_score_col or cell.column in failure_cols
                ):
                    cell.fill = zero_fill
                if cell.value is not None:
                    w = _str_display_width(str(cell.value))
                    col_letter = cell.column_letter
                    col_widths[col_letter] = max(col_widths.get(col_letter, 0), w)

        for col_letter, w in col_widths.items():
            ws.column_dimensions[col_letter].width = min(w + 2, 60)  # 最多60宽，留2字符边距

    # ── 标题行超链接（内部链接须用 Hyperlink(location=...)）─────────────
    for sheet_name in [
        "实盘候选评分",
        "模型Alpha评分",
        "模型Seed稳定性",
        "交易参数收益评分",
        "实验对比",
    ]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            metric_name = str(cell.value) if cell.value else ""
            if metric_name in desc_row_map:
                target_row = desc_row_map[metric_name]
                cell.hyperlink = Hyperlink(
                    ref=cell.coordinate,
                    location=f"'指标说明'!A{target_row}",
                )
                cell.font = font_link


def print_comparison_table(df: pd.DataFrame) -> None:
    """控制台打印可读的对比表（精简版）"""
    if df.empty:
        logger.info("对比表为空")
        return

    display_cols = [
        COL_NAMES["wf_run_id"],
        "重点Top20命中率均值",
        "重点Top20收益中位数均值",
        "重点Top30命中率均值",
        "重点Top30收益中位数均值",
        "综合得分",
        "选股综合得分",
        COL_NAMES["n_splits"],
        COL_NAMES["model_version_range"],
        COL_NAMES["daily_rankic_mean"],
        COL_NAMES["icir"],
        COL_NAMES["oos_top30_lift_mean"],
        COL_NAMES["chain_cagr"],
        COL_NAMES["chain_max_drawdown"],
        COL_NAMES["chain_total_return"],
        COL_NAMES["oos_cross_split_ir"],
        COL_NAMES["oos_rankic_ir_mean"],
        COL_NAMES["oos_top30_win_rate"],
        COL_NAMES["oos_top30_median_mean"],
        COL_NAMES["oos_top30_worst_median"],
        COL_NAMES["bt_annual_return_mean"],
        COL_NAMES["bt_sharpe_mean"],
        COL_NAMES["bt_max_drawdown_worst"],
        COL_NAMES["bt_win_rate"],
        COL_NAMES["val_rankic_ir_mean"],
        COL_NAMES["train_val_ir_gap"],
        COL_NAMES["best_iter_mean"],
        COL_NAMES["label_column"],
        COL_NAMES["task"],
        COL_NAMES["n_estimators"],
        COL_NAMES["max_depth"],
        COL_NAMES["learning_rate"],
    ]
    show_cols = [c for c in display_cols if c in df.columns]
    pd.set_option("display.max_columns", None)
    pd.set_option("display.width", 200)
    pd.set_option("display.float_format", "{:.4f}".format)
    logger.info("\n" + df[show_cols].to_string(index=True))


def write_empty_report(output_path: Path, source_label: str) -> None:
    """为无数据来源生成占位 Excel，保证固定输出文件存在。"""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    placeholder_df = pd.DataFrame(
        {
            "状态": ["无可用数据"],
            "来源": [source_label],
            "说明": ["当前来源目录下未找到 walk_forward_summary_*.csv"],
        }
    )
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        placeholder_df.to_excel(writer, sheet_name="实验对比", index=False)


def _score_sheet_or_placeholder(df: pd.DataFrame, sheet_name: str, reason: str) -> pd.DataFrame:
    """评分表为空时仍输出占位说明，保持 Excel 工作表结构稳定。"""
    if not df.empty:
        return df
    return pd.DataFrame(
        {
            "状态": ["无可用评分"],
            "评分视角": [sheet_name],
            "说明": [reason],
        }
    )


def generate_comparison_report(
    raw_dirs: list[Path],
    output_path: Path,
    source_label: str,
    data_root: Optional[Path] = None,
    write_empty_output: bool = False,
) -> bool:
    """加载指定来源的汇总CSV并写出对比 Excel。"""
    logger.info("-" * 70)
    logger.info(f"来源标签:   {source_label}")
    if len(raw_dirs) == 1:
        logger.info(f"汇总CSV目录: {raw_dirs[0]}")
    else:
        logger.info(f"汇总CSV目录: 共 {len(raw_dirs)} 个 raw 目录（来源: {source_label}）")
    logger.info(f"输出路径:     {output_path}")

    all_df = load_all_summaries_from_raw_dirs(raw_dirs, data_root=data_root)
    if all_df.empty:
        logger.warning(f"[{source_label}] 没有可用数据，跳过")
        if write_empty_output:
            write_empty_report(output_path, source_label)
            logger.info(f"[{source_label}] 已生成空白占位文件: {output_path}")
            return True
        return False

    comp_df = build_comparison_table(all_df)
    if comp_df.empty:
        logger.warning(f"[{source_label}] 构建对比表失败，跳过")
        return False

    comp_df.insert(1, "综合得分", compute_composite_score(comp_df))
    comp_df.insert(2, "选股综合得分", compute_selection_score(comp_df))
    comp_df = reorder_comparison_columns(comp_df)
    logger.info(
        f"综合得分计算完成（参与评分指标数: {sum(1 for k, _, _ in SCORE_CONFIG if COL_NAMES.get(k) in comp_df.columns)}）"
    )
    logger.info("选股综合得分计算完成（指标: RankIC均值30%/ICIR30%/Top30超额40%）")

    desc_df = build_metric_descriptions()
    split_df = build_split_detail_table(all_df)
    logger.info(f"逐Split明细表: {len(split_df)} 行")
    period_stability_df = build_period_stability_table(comp_df)
    logger.info(f"跨时间段稳定性表: {len(period_stability_df)} 行")
    model_alpha_df = build_model_alpha_score_table(comp_df)
    logger.info(f"模型Alpha评分表: {len(model_alpha_df)} 行")
    model_seed_stability_df = build_model_seed_stability_table(comp_df, model_alpha_df)
    logger.info(f"模型Seed稳定性表: {len(model_seed_stability_df)} 行")
    trade_score_df = build_trade_param_score_table(comp_df)
    logger.info(f"交易参数收益评分表: {len(trade_score_df)} 行")
    candidate_df = build_live_candidate_score_table(comp_df, model_alpha_df, trade_score_df)
    logger.info(f"实盘候选评分表: {len(candidate_df)} 行")
    comp_df = sort_by_run_time(comp_df)
    display_comp_df = compact_experiment_sheet_for_display(comp_df, source_label)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _score_sheet_or_placeholder(
            candidate_df,
            "实盘候选评分",
            "缺少模型Alpha评分或有效交易参数配对评分，无法计算实盘候选分。",
        ).to_excel(writer, sheet_name="实盘候选评分", index=False)
        _score_sheet_or_placeholder(
            model_alpha_df,
            "模型Alpha评分",
            "缺少可聚合的模型参数或选股指标，无法计算模型Alpha分。",
        ).to_excel(writer, sheet_name="模型Alpha评分", index=False)
        _score_sheet_or_placeholder(
            model_seed_stability_df,
            "模型Seed稳定性",
            "缺少 seed 字段或模型Alpha评分，无法按忽略 seed 的超参分组统计稳定性。",
        ).to_excel(writer, sheet_name="模型Seed稳定性", index=False)
        _score_sheet_or_placeholder(
            trade_score_df,
            "交易参数收益评分",
            "没有相同模型参数 + 相同时间段下的多交易参数候选，无法计算配对交易参数评分。",
        ).to_excel(writer, sheet_name="交易参数收益评分", index=False)
        display_comp_df.to_excel(writer, sheet_name="实验对比", index=False)
        if not period_stability_df.empty:
            period_stability_df.to_excel(writer, sheet_name="跨时间段稳定性", index=False)
        desc_df.to_excel(writer, sheet_name="指标说明", index=False)
        if not split_df.empty:
            split_df.to_excel(writer, sheet_name="逐Split明细", index=False)
        format_excel_output(writer.book, desc_df)
    logger.info(f"[{source_label}] 对比表已保存: {output_path}（{len(comp_df)} 个实验）")
    return True


def run_auto_compare_jobs(data_root: Path) -> list[Path]:
    """无参模式：自动扫描 raw 与 batches 两类来源。"""
    output_paths: list[Path] = []
    jobs = build_auto_compare_jobs(data_root)
    logger.info("无参模式：自动扫描 raw 与 batches 目录")
    for job in jobs:
        if generate_comparison_report(
            job["raw_dirs"],
            job["output_path"],
            job["label"],
            data_root=data_root,
            write_empty_output=True,
        ):
            output_paths.append(job["output_path"])
    return output_paths


def main():
    parser = argparse.ArgumentParser(description="Walk-forward 实验对比分析")
    parser.add_argument(
        "--data-root",
        type=str,
        default=None,
        help="数据根目录；未指定时使用 configs/base.yaml 中的 data.root",
    )
    parser.add_argument(
        "--raw-dir",
        type=str,
        default=None,
        help="walk_forward 汇总CSV目录，默认 {data_root}/walk_forward/raw",
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="对比Excel输出路径，默认 {data_root}/walk_forward/wf_comparison.xlsx",
    )
    args = parser.parse_args()

    setup_logger()

    effective_data_root = Path(args.data_root or get_data_root())

    logger.info("=" * 70)
    logger.info("Walk-forward 实验对比分析")
    logger.info("=" * 70)

    if args.raw_dir is None and args.output is None:
        output_paths = run_auto_compare_jobs(effective_data_root)
        if len(output_paths) == 0:
            logger.error("raw 与 batches 均没有可用数据，退出")
            return
        logger.info(f"自动扫描完成，共生成 {len(output_paths)} 个对比文件")
    else:
        raw_dir = (
            Path(args.raw_dir) if args.raw_dir else effective_data_root / "walk_forward" / "raw"
        )
        output_path = (
            Path(args.output)
            if args.output
            else effective_data_root / "walk_forward" / "wf_comparison.xlsx"
        )
        if not generate_comparison_report(
            [raw_dir], output_path, "single", data_root=effective_data_root
        ):
            logger.error("没有可用数据，退出")
            return

    logger.info("=" * 70)


if __name__ == "__main__":
    main()
