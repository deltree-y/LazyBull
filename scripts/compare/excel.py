# -*- coding: utf-8 -*-
"""openpyxl Excel 格式化与控制台打印。"""

from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.hyperlink import Hyperlink
from loguru import logger
import pandas as pd

import sys
from pathlib import Path

project_root = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(project_root))

from scripts.compare.constants import (
    CANDIDATE_SCORE_CONFIG,
    COL_NAMES,
    MODEL_ALPHA_SCORE_CONFIG,
    SCORE_CONFIG,
    TRADE_YIELD_SCORE_CONFIG,
)
from scripts.compare.detail_display import _str_display_width


def _weight_to_green_fill(weight: float, max_weight: float) -> PatternFill:
    """将权重映射为浅绿→中绿的填充色（权重越高越深，黑字始终可读）

    颜色范围：
      最小权重 → RGB(220, 245, 220)  极浅绿
      最大权重 → RGB(130, 215, 130)  中等绿（与黑字对比度 ≈ 8:1，远超 WCAG AA 4.5:1）
    """
    t = min(weight / max_weight, 1.0) if max_weight > 0 else 0.0
    r = int(round(220 - t * 90))  # 220 → 130
    g = int(round(245 - t * 30))  # 245 → 215
    b = int(round(220 - t * 90))  # 220 → 130
    return PatternFill(fill_type="solid", fgColor=f"{r:02X}{g:02X}{b:02X}")


def _weight_to_palette_fill(weight: float, max_weight: float, palette: str) -> PatternFill:
    """将权重映射为指定色系的浅色填充，权重越高颜色越深。"""
    if palette == "blue":
        start, end = (225, 239, 255), (132, 181, 232)
    elif palette == "orange":
        start, end = (255, 235, 205), (242, 174, 88)
    elif palette == "red":
        start, end = (255, 224, 224), (242, 150, 150)
    else:
        start, end = (220, 245, 220), (130, 215, 130)
    t = min(weight / max_weight, 1.0) if max_weight > 0 else 0.0
    r = int(round(start[0] + (end[0] - start[0]) * t))
    g = int(round(start[1] + (end[1] - start[1]) * t))
    b = int(round(start[2] + (end[2] - start[2]) * t))
    return PatternFill(fill_type="solid", fgColor=f"{r:02X}{g:02X}{b:02X}")


def _score_column_fills() -> dict[str, dict[str, PatternFill]]:
    """构建各评分 sheet 的列填充配置。"""
    max_model = max(w for _, w, _ in MODEL_ALPHA_SCORE_CONFIG)
    max_trade = max(w for _, w, _ in TRADE_YIELD_SCORE_CONFIG)
    max_candidate = max(w for _, w, _ in CANDIDATE_SCORE_CONFIG)
    return {
        "模型Alpha评分": {
            "模型Alpha分": _weight_to_palette_fill(max_model, max_model, "blue"),
            **{
                col: _weight_to_palette_fill(weight, max_model, "blue")
                for col, weight, _ in MODEL_ALPHA_SCORE_CONFIG
            },
        },
        "模型Seed稳定性": {
            "Seed稳健分": _weight_to_palette_fill(max_model, max_model, "blue"),
            "模型Alpha分均值": _weight_to_palette_fill(0.30, max_model, "blue"),
            "模型Alpha分标准差": _weight_to_palette_fill(0.15, max_model, "blue"),
            "模型Alpha分最差": _weight_to_palette_fill(0.25, max_model, "blue"),
            "模型Alpha分最好": _weight_to_palette_fill(0.10, max_model, "blue"),
        },
        "交易参数收益评分": {
            "交易收益分": _weight_to_palette_fill(max_trade, max_trade, "orange"),
            "交易稳健分": _weight_to_palette_fill(max_trade, max_trade, "orange"),
            **{
                col: _weight_to_palette_fill(weight, max_trade, "orange")
                for col, weight, _ in TRADE_YIELD_SCORE_CONFIG
            },
        },
        "实盘候选评分": {
            "实盘候选分": _weight_to_palette_fill(max_candidate, max_candidate, "green"),
            "实盘候选原始分": _weight_to_palette_fill(max_candidate, max_candidate, "green"),
            **{
                col: _weight_to_palette_fill(weight, max_candidate, "green")
                for col, weight, _ in CANDIDATE_SCORE_CONFIG
            },
        },
    }


def format_excel_output(wb, desc_df: pd.DataFrame) -> None:
    """对 Excel 工作簿应用格式化

    - 全局字体: 微软雅黑 9 号
    - 冻结标题行（两个 sheet 均适用）
    - 实验对比标题行超链接跳转至指标说明对应行
    - 自动列宽（CJK 双倍宽）
    - 参与综合得分的列着浅绿背景，权重越高越深
    """
    # 构建 指标名 → 指标说明 sheet 行号的映射（第 1 行为标题，数据从第 2 行起）
    desc_row_map: dict[str, int] = {}
    for i, row in desc_df.iterrows():
        desc_row_map[str(row["指标名"])] = int(i) + 2  # +2: header占第1行，数据从第2行

    font_normal = Font(name="微软雅黑", size=9)
    font_link = Font(name="微软雅黑", size=9, color="0563C1", underline="single")

    # ── 构建 中文列名 → 填充 的映射（用于各评分 sheet）─────────────────
    _max_w = max(w for _, w, _ in SCORE_CONFIG) if SCORE_CONFIG else 1.0
    # 综合得分列本身也着色，用最深绿（权重等同最大权重）
    score_cn_fills: dict[str, PatternFill] = {
        "综合得分": _weight_to_green_fill(_max_w, _max_w),
    }
    for eng_key, weight, _ in SCORE_CONFIG:
        col_cn = COL_NAMES.get(eng_key)
        if col_cn:
            score_cn_fills[col_cn] = _weight_to_green_fill(weight, _max_w)

    sheet_score_fills = _score_column_fills()
    key_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    zero_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")
    score_header_fill = PatternFill(fill_type="solid", fgColor="274E13")
    score_header_font = Font(name="微软雅黑", size=9, color="FFFFFF", bold=True)
    key_cols_cn = {
        "重点说明",
        "重点Top20最新名单",
        "重点Top30最新名单",
        "重点Top20命中率均值",
        "重点Top20收益中位数均值",
        "重点Top30命中率均值",
        "重点Top30收益中位数均值",
    }

    sheet_col_letter_fill: dict[str, dict[str, PatternFill]] = {}
    for sheet_name in wb.sheetnames:
        ws_ref = wb[sheet_name]
        col_letter_fill: dict[str, PatternFill] = {}
        local_fills = (
            score_cn_fills if sheet_name == "实验对比" else sheet_score_fills.get(sheet_name, {})
        )
        for cell in next(ws_ref.iter_rows(min_row=1, max_row=1)):
            if cell.value and str(cell.value) in local_fills:
                col_letter_fill[cell.column_letter] = local_fills[str(cell.value)]
            elif sheet_name == "实验对比" and cell.value and str(cell.value) in key_cols_cn:
                col_letter_fill[cell.column_letter] = key_fill
        sheet_col_letter_fill[sheet_name] = col_letter_fill

    # ── 全局字体、冻结、列宽、绿色背景 ──────────────────────────────────
    all_sheets = [
        s
        for s in [
            "实盘候选评分",
            "模型Alpha评分",
            "模型Seed稳定性",
            "交易参数收益评分",
            "实验对比",
            "跨时间段稳定性",
            "指标说明",
            "逐Split明细",
        ]
        if s in wb.sheetnames
    ]
    for sheet_name in all_sheets:
        ws = wb[sheet_name]
        ws.freeze_panes = "A2"
        header_values = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        zero_score_col = None
        failure_cols = set()
        if sheet_name == "实盘候选评分":
            for idx, value in enumerate(header_values, 1):
                if value == "实盘候选分":
                    zero_score_col = idx
                if value in {
                    "候选门槛失败原因",
                    "候选门槛通过",
                    "模型Alpha门槛通过",
                    "有效配对门槛通过",
                    "最大回撤门槛通过",
                    "最差CAGR门槛通过",
                }:
                    failure_cols.add(idx)

        col_widths: dict[str, int] = {}
        for row in ws.iter_rows():
            is_zero_candidate = False
            if sheet_name == "实盘候选评分" and row[0].row > 1 and zero_score_col is not None:
                score_value = row[zero_score_col - 1].value
                try:
                    is_zero_candidate = float(score_value) == 0.0
                except (TypeError, ValueError):
                    is_zero_candidate = False
            for cell in row:
                cell.font = font_normal
                if cell.row == 1 and cell.column_letter in sheet_col_letter_fill.get(
                    sheet_name, {}
                ):
                    cell.fill = score_header_fill
                    cell.font = score_header_font
                elif cell.column_letter in sheet_col_letter_fill.get(sheet_name, {}):
                    cell.fill = sheet_col_letter_fill[sheet_name][cell.column_letter]
                if is_zero_candidate and (
                    cell.column == zero_score_col or cell.column in failure_cols
                ):
                    cell.fill = zero_fill
                if cell.value is not None:
                    w = _str_display_width(str(cell.value))
                    col_letter = cell.column_letter
                    col_widths[col_letter] = max(col_widths.get(col_letter, 0), w)

        for col_letter, w in col_widths.items():
            ws.column_dimensions[col_letter].width = min(w + 2, 60)  # 最多60宽，留2字符边距

    # ── 标题行超链接（内部链接须用 Hyperlink(location=...)）─────────────
    for sheet_name in [
        "实盘候选评分",
        "模型Alpha评分",
        "模型Seed稳定性",
        "交易参数收益评分",
        "实验对比",
    ]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            metric_name = str(cell.value) if cell.value else ""
            if metric_name in desc_row_map:
                target_row = desc_row_map[metric_name]
                cell.hyperlink = Hyperlink(
                    ref=cell.coordinate,
                    location=f"'指标说明'!A{target_row}",
                )
                cell.font = font_link


def print_comparison_table(df: pd.DataFrame) -> None:
    """控制台打印可读的对比表（精简版）"""
    if df.empty:
        logger.info("对比表为空")
        return

    display_cols = [
        COL_NAMES["wf_run_id"],
        "重点Top20命中率均值",
        "重点Top20收益中位数均值",
        "重点Top30命中率均值",
        "重点Top30收益中位数均值",
        "综合得分",
        "选股综合得分",
        COL_NAMES["n_splits"],
        COL_NAMES["model_version_range"],
        COL_NAMES["daily_rankic_mean"],
        COL_NAMES["icir"],
        COL_NAMES["oos_top30_lift_mean"],
        COL_NAMES["chain_cagr"],
        COL_NAMES["chain_max_drawdown"],
        COL_NAMES["chain_total_return"],
        COL_NAMES["oos_cross_split_ir"],
        COL_NAMES["oos_rankic_ir_mean"],
        COL_NAMES["oos_top30_win_rate"],
        COL_NAMES["oos_top30_median_mean"],
        COL_NAMES["oos_top30_worst_median"],
        COL_NAMES["bt_annual_return_mean"],
        COL_NAMES["bt_sharpe_mean"],
        COL_NAMES["bt_max_drawdown_worst"],
        COL_NAMES["bt_win_rate"],
        COL_NAMES["val_rankic_ir_mean"],
        COL_NAMES["train_val_ir_gap"],
        COL_NAMES["best_iter_mean"],
        COL_NAMES["label_column"],
        COL_NAMES["task"],
        COL_NAMES["n_estimators"],
        COL_NAMES["max_depth"],
        COL_NAMES["learning_rate"],
    ]
    show_cols = [c for c in display_cols if c in df.columns]
    pd.set_option("display.max_columns", None)
    pd.set_option("display.width", 200)
    pd.set_option("display.float_format", "{:.4f}".format)
    logger.info("\n" + df[show_cols].to_string(index=True))


def write_empty_report(output_path: Path, source_label: str) -> None:
    """为无数据来源生成占位 Excel，保证固定输出文件存在。"""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    placeholder_df = pd.DataFrame(
        {
            "状态": ["无可用数据"],
            "来源": [source_label],
            "说明": ["当前来源目录下未找到 walk_forward_summary_*.csv"],
        }
    )
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        placeholder_df.to_excel(writer, sheet_name="实验对比", index=False)


def _score_sheet_or_placeholder(df: pd.DataFrame, sheet_name: str, reason: str) -> pd.DataFrame:
    """评分表为空时仍输出占位说明，保持 Excel 工作表结构稳定。"""
    if not df.empty:
        return df
    return pd.DataFrame(
        {
            "状态": ["无可用评分"],
            "评分视角": [sheet_name],
            "说明": [reason],
        }
    )
